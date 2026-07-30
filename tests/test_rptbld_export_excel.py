"""
export_report_to_excel: the existing flat-mode branch (Overview + "Report
Data" sheets) must stay byte-identical for templates that don't use
row_groups. Phase 3 only adds a new, additive "Report Data (Pivot)" sheet
when row_groups is configured, built from pivot_report_data() (phase 2).
Every assertion here reads the generated workbook back with openpyxl and
checks actual cell values/fills -- not just "no exception was raised."

The one exception is the cross-site composer template (see
TestCrossSiteTemplateExport below): its flat "Report Data" sheet is skipped
entirely (redundant -- the reference workbook it replicates is a single data
sheet), and its grouped sheet is named "Sheet1" instead of the generic
"Report Data (Pivot)". Every other template keeps the 3-sheet shape above.
"""
import io

import openpyxl
import pytest

from app.modules.RPTBLD.service import create_report_template, export_report_to_excel


class TestFlatModeExportUnaffected:
    def test_no_row_groups_produces_flat_only_workbook(
        self, make_site, make_form, make_field, make_workflow, make_reporting_period,
        make_submission, make_submission_value, make_user, make_access_grant, created_objects,
    ):
        site = make_site()
        form, form_version = make_form()
        field, fv = make_field(form, form_version, "cargo_code")
        workflow_version = make_workflow([])
        period = make_reporting_period(site, year=2026, month=3)
        sub = make_submission(
            site, form, form_version, period, workflow_version,
            status="Approved", is_locked=True,
        )
        make_submission_value(sub, field, fv, raw_value="42")

        user = make_user()
        make_access_grant(user, "report", scope_type="global", can_export=True, can_view=True)

        template = create_report_template(
            "Flat Only Report", f"test-flat-xlsx-{user.id}", None, "global", None,
            {"form_ids": [form.id], "site_ids": [site.id]}, user.id,
        )
        created_objects.append(template)

        raw = export_report_to_excel(template.id, user.id)
        wb = openpyxl.load_workbook(io.BytesIO(raw))

        assert set(wb.sheetnames) == {"Overview", "Report Data"}

        ws = wb["Report Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers == ["Period", "Site Name", "Form Name", "Field Code", "Field Name", "Field Type", "Value", "Unit"]
        assert ws.cell(row=2, column=4).value == "cargo_code"
        assert ws.cell(row=2, column=7).value == 42.0


class TestPivotModeExport:
    def _setup(
        self, make_site, make_form, make_field, make_workflow, make_reporting_period,
        make_submission, make_submission_value, make_user, make_access_grant, created_objects,
    ):
        from app.modules.FRMULA.model import FormulaVersion
        from app.modules.FRMULA.service import create_formula, publish_formula_version

        site = make_site()
        form, form_version = make_form()
        cargo_field, cargo_fv = make_field(form, form_version, "cargo_code")
        scope1_field, scope1_fv = make_field(form, form_version, "scope1_code")  # aliased, never submitted -> blank cell
        power_field, power_fv = make_field(form, form_version, "power_code")
        diesel_field, diesel_fv = make_field(form, form_version, "diesel_code")

        workflow_version = make_workflow([])
        period = make_reporting_period(site, year=2026, month=3)
        sub = make_submission(
            site, form, form_version, period, workflow_version,
            status="Approved", is_locked=True,
        )
        make_submission_value(sub, cargo_field, cargo_fv, raw_value="1000")
        make_submission_value(sub, power_field, power_fv, raw_value="250")
        make_submission_value(sub, diesel_field, diesel_fv, raw_value="180")
        # scope1_field deliberately has no SubmissionValue -- the aliased
        # column exists but this row's value is missing.

        user = make_user()
        make_access_grant(user, "report", scope_type="global", can_export=True, can_view=True)

        # energy_fossil is never aliased anywhere in this config -- a
        # genuinely-missing metric, same shape as a site with no petrol/IFO
        # consumption. This column must render blank, not error.
        n_formula = create_formula(
            "Blank Col", f"test-xlsx-n-{user.id}", "cargo + energy_fossil",
            {"cargo": {}, "energy_fossil": {}}, user.id, context="report",
        )
        created_objects.append(n_formula)
        n_version = FormulaVersion.query.filter_by(formula_id=n_formula.id, version_number=1).one()
        created_objects.append(n_version)
        publish_formula_version(n_version.id, user.id)

        config_json = {
            "form_ids": [form.id],
            "site_ids": [site.id],
            "row_groups": [
                {
                    "id": "core", "label": "Core", "subtotal_label": "Total",
                    "site_ids": [site.id],
                    "include_in_grand_total": True, "is_reference_base": True,
                },
            ],
            "metric_aliases": {
                "cargo": [{"site_id": site.id, "field_ids": [cargo_field.id], "op": "single", "verified": True}],
                "scope1": [{"site_id": site.id, "field_ids": [scope1_field.id], "op": "single", "verified": True}],
                "power_specific": [{"site_id": site.id, "field_ids": [power_field.id], "op": "single", "verified": False}],
                "diesel_specific": [{"site_id": site.id, "field_ids": [diesel_field.id], "op": "single", "verified": True}],
            },
            "computed_columns": [
                {"id": "N", "label": "Blank Col", "formula_id": n_formula.id},
            ],
        }

        template = create_report_template(
            "Pivot Export Report", f"test-xlsx-pivot-{user.id}", None, "global", None, config_json, user.id,
        )
        created_objects.append(template)
        return template, user, site

    def test_pivot_export_produces_both_sheets(
        self, make_site, make_form, make_field, make_workflow, make_reporting_period,
        make_submission, make_submission_value, make_user, make_access_grant, created_objects,
    ):
        template, user, site = self._setup(
            make_site, make_form, make_field, make_workflow, make_reporting_period,
            make_submission, make_submission_value, make_user, make_access_grant, created_objects,
        )

        raw = export_report_to_excel(template.id, user.id)
        wb = openpyxl.load_workbook(io.BytesIO(raw))

        assert {"Overview", "Report Data", "Report Data (Pivot)"} <= set(wb.sheetnames)
        # The flat sheet's own headers are still exactly what they were before this phase.
        ws_flat = wb["Report Data"]
        headers = [ws_flat.cell(row=1, column=c).value for c in range(1, 9)]
        assert headers == ["Period", "Site Name", "Form Name", "Field Code", "Field Name", "Field Type", "Value", "Unit"]

    def test_pivot_sheet_layout_and_values(
        self, make_site, make_form, make_field, make_workflow, make_reporting_period,
        make_submission, make_submission_value, make_user, make_access_grant, created_objects,
    ):
        template, user, site = self._setup(
            make_site, make_form, make_field, make_workflow, make_reporting_period,
            make_submission, make_submission_value, make_user, make_access_grant, created_objects,
        )

        raw = export_report_to_excel(template.id, user.id)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb["Report Data (Pivot)"]

        # Columns: 1=Row Group, 2=Site, 3=Cargo, 4=Scope1, 5=Power Specific,
        # 6=Diesel Specific, 7=Blank Col (computed). energy_elec/energy_fossil/
        # scope2/total_ghg/petrol_specific/ifo_specific never aliased -> no columns.
        headers = [ws.cell(row=2, column=c).value for c in range(1, 8)]
        assert headers == ["Row Group", "Site", "Cargo", "Scope1", "Power Specific", "Diesel Specific", "Blank Col"]

        # Banner merges over exactly the 2 specific-consumption columns
        # actually present (columns 5-6), not a hardcoded fixed range.
        assert ws.cell(row=1, column=5).value == "Specific Consumptions"
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        assert "E1:F1" in merged_ranges
        assert ws.cell(row=1, column=3).value is None  # no banner over Cargo
        assert ws.cell(row=1, column=7).value is None  # no banner over the computed column

        # Row 3 is the (only) site row.
        assert ws.cell(row=3, column=1).value == "Core"
        assert ws.cell(row=3, column=3).value == 1000.0   # cargo, verified
        assert ws.cell(row=3, column=4).value is None     # scope1 aliased but never submitted -> blank, not 0/"None"
        assert ws.cell(row=3, column=5).value == 250.0    # power_specific, unverified
        assert ws.cell(row=3, column=6).value == 180.0    # diesel_specific, verified

        # Unverified sourced cell is amber-flagged; verified ones are not.
        amber = "FEF3C7"
        assert ws.cell(row=3, column=5).fill.start_color.rgb[-6:] == amber
        assert ws.cell(row=3, column=3).fill.start_color.rgb in (None, "00000000")
        assert ws.cell(row=3, column=6).fill.start_color.rgb in (None, "00000000")

        # A computed column referencing a never-aliased metric (energy_fossil)
        # renders as a genuinely blank cell -- not an "Unknown formula
        # variable" error, and not 0 or some other placeholder.
        assert ws.cell(row=3, column=7).value is None

        # Row 4: subtotal, bold, labeled from the group's subtotal_label.
        assert ws.cell(row=4, column=2).value == "Total"
        assert ws.cell(row=4, column=2).font.bold is True
        assert ws.cell(row=4, column=3).value == 1000.0

        # Row 5: grand total, bold, default label since config_json carries none.
        assert ws.cell(row=5, column=2).value == "All Locations"
        assert ws.cell(row=5, column=2).font.bold is True
        assert ws.cell(row=5, column=3).value == 1000.0


class TestCrossSiteTemplateExport:
    """export_report_to_excel used to route EVERY row_groups-configured
    template (including this one) through pivot_report_data, whose
    _evaluate_computed_columns unconditionally reads col["formula_id"] --
    a KeyError for this template's kind="cross_site" computed_columns
    entries (they carry no formula_id by design). Confirmed via a live
    traceback before fixing: export_report_to_excel must detect this
    template by code and build the pivot sheet from
    compose_cross_site_intensity_report instead."""

    def _setup(
        self, make_site, make_form, make_field, make_formula_version, make_workflow,
        make_user, make_workbook, make_access_grant, make_reporting_period,
        make_submission, make_submission_value, created_objects,
    ):
        from app.modules.FRMULA.model import FormulaVersion
        from app.modules.FRMULA.service import create_formula, publish_formula_version
        from app.modules.RPTBLD.service import CROSS_SITE_COMPOSER_TEMPLATE_CODE

        spt_site = make_site()
        non_spt_site = make_site()

        form, form_version = make_form()
        cargo_field, cargo_fv = make_field(form, form_version, "cargo_qty", frequency="monthly")
        scope1_field, scope1_fv = make_field(form, form_version, "scope1_qty", frequency="monthly")
        scope2_field, scope2_fv = make_field(form, form_version, "scope2_qty", frequency="monthly")

        cargo_formula = make_formula_version("SUM_MONTHS(cargo_qty)", {"cargo_qty": {}})
        scope1_formula = make_formula_version("SUM_MONTHS(scope1_qty)", {"scope1_qty": {}})
        scope2_formula = make_formula_version("SUM_MONTHS(scope2_qty)", {"scope2_qty": {}})
        fy_cargo, fy_cargo_fv = make_field(
            form, form_version, "fy_total_cargo", field_type="calculated", frequency="annual",
            field_config={"formula_version_id": cargo_formula.id, "display_region": "below_monthly_table"},
        )
        fy_scope1, fy_scope1_fv = make_field(
            form, form_version, "fy_total_scope1", field_type="calculated", frequency="annual",
            field_config={"formula_version_id": scope1_formula.id, "display_region": "below_monthly_table"},
        )
        fy_scope2, fy_scope2_fv = make_field(
            form, form_version, "fy_total_scope2", field_type="calculated", frequency="annual",
            field_config={"formula_version_id": scope2_formula.id, "display_region": "below_monthly_table"},
        )

        for site, cargo_val, scope1_val, scope2_val in (
            (spt_site, 1000, 100, 50),
            (non_spt_site, 500, 20, 10),
        ):
            user = make_user()
            workflow_version = make_workflow([user])
            make_workbook(form, site, workflow_version=workflow_version, submitters=[user])
            make_access_grant(
                user, "submission", scope_type="global",
                can_view=True, can_create=True, can_edit=True, can_submit=True,
            )
            for month in (4, 5, 6, 7, 8, 9, 10, 11, 12, 1, 2, 3):
                year = 2026 if month >= 4 else 2027
                period = make_reporting_period(site, year=year, month=month)
                submission = make_submission(
                    site, form, form_version, period, workflow_version,
                    status="Approved", is_locked=True,
                )
                make_submission_value(submission, cargo_field, cargo_fv, raw_value=str(cargo_val // 12))
                make_submission_value(submission, scope1_field, scope1_fv, raw_value=str(scope1_val // 12))
                make_submission_value(submission, scope2_field, scope2_fv, raw_value=str(scope2_val // 12))

        formula_user = make_user()
        m_formula = create_formula(
            "Total GHG", f"test-xlsx-crosssite-m-{formula_user.id}", "scope1 + scope2",
            {"scope1": {}, "scope2": {}}, formula_user.id, context="report",
        )
        created_objects.append(m_formula)
        m_version = FormulaVersion.query.filter_by(formula_id=m_formula.id, version_number=1).one()
        created_objects.append(m_version)
        publish_formula_version(m_version.id, formula_user.id)

        export_user = make_user()
        make_access_grant(export_user, "report", scope_type="global", can_export=True, can_view=True)

        def alias_for(field_key):
            return [
                {"site_id": s.id, "field_ids": [f.id], "op": "single", "verified": True}
                for s, f in ((spt_site, {"cargo": fy_cargo, "scope1": fy_scope1, "scope2": fy_scope2}[field_key]),
                             (non_spt_site, {"cargo": fy_cargo, "scope1": fy_scope1, "scope2": fy_scope2}[field_key]))
            ]

        config_json = {
            "row_groups": [
                {
                    "id": "spt", "label": "SPT", "subtotal_label": "Total",
                    "site_ids": [spt_site.id], "is_reference_base": True, "include_in_grand_total": True,
                },
                {
                    "id": "non_spt", "label": "Non-SPT", "subtotal_label": "Total All Locations (incl. Non SPT)",
                    "site_ids": [non_spt_site.id], "is_reference_base": False, "include_in_grand_total": True,
                    "suppress_own_subtotal": True,
                },
            ],
            "metric_aliases": {
                "cargo": alias_for("cargo"),
                "scope1": alias_for("scope1"),
                "scope2": alias_for("scope2"),
            },
            "computed_columns": [
                {"id": "total_ghg_emission", "label": "Total GHG Emission", "formula_id": m_formula.id},
                {"id": "pct_contribution_total_ghg", "label": "% Contribution", "kind": "cross_site"},
                {"id": "variation_from_avg_intensity", "label": "Variation", "kind": "cross_site"},
            ],
            "include_unapproved": True,
            "grand_total_label": "Total All Locations (incl. Non SPT)",
        }

        template = create_report_template(
            "Cross-Site Export Test", CROSS_SITE_COMPOSER_TEMPLATE_CODE, None, "global", None, config_json, export_user.id,
        )
        created_objects.append(template)

        return template, export_user

    def test_export_does_not_raise_and_suppresses_non_spt_subtotal(
        self, make_site, make_form, make_field, make_formula_version, make_workflow,
        make_user, make_workbook, make_access_grant, make_reporting_period,
        make_submission, make_submission_value, created_objects,
    ):
        template, export_user = self._setup(
            make_site, make_form, make_field, make_formula_version, make_workflow,
            make_user, make_workbook, make_access_grant, make_reporting_period,
            make_submission, make_submission_value, created_objects,
        )

        raw = export_report_to_excel(template.id, export_user.id)  # must not raise KeyError('formula_id')
        wb = openpyxl.load_workbook(io.BytesIO(raw))

        # Exactly 2 sheets for this template -- Overview + its own grouped
        # data sheet named "Sheet1" (mirroring the reference workbook it
        # replicates) -- no separate, redundant flat "Report Data" dump.
        assert set(wb.sheetnames) == {"Overview", "Sheet1"}
        ws = wb["Sheet1"]

        # Column A header is "Location" (a single merged column, not the
        # generic "Row Group" + "Site" pair) -- matches the reference
        # workbook, which has no separate per-row group column.
        assert ws.cell(row=2, column=1).value == "Location"

        # The real Sheet1 header order/labels, filtered down to what this
        # minimal fixture actually configures (cargo/scope1/scope2 raw +
        # total_ghg_emission/pct_contribution/variation computed) -- energy
        # columns are absent here since energy_elec/energy_fossil were never
        # aliased in this test's config, not because the ordering is wrong.
        headers = [ws.cell(row=2, column=c).value for c in range(2, ws.max_column + 1)]
        assert headers == [
            "Cargo Handled (MT)",
            "GHG Emission Scope-1 (tCO2e)",
            "GHG Emission Scope-2 (tCO2e)",
            "Total GHG Emission (tCO2e)",
            "% Contribution of Total Emissions",
            "Variation from Average Intensity",
        ]

        location_column = []
        for r in range(3, ws.max_row + 1):
            location_column.append(ws.cell(row=r, column=1).value)

        # 1 SPT site row, its own "Total" subtotal, 1 Non-SPT site row, then
        # ONLY the grand total row -- the Non-SPT group's own subtotal must
        # not appear as a separate (mislabeled, wrong-value) row.
        spt_site, non_spt_site = template.config_json["row_groups"][0]["site_ids"][0], template.config_json["row_groups"][1]["site_ids"][0]
        from app.modules.SITEMST.model import Site
        spt_name = Site.query.get(spt_site).name
        non_spt_name = Site.query.get(non_spt_site).name
        assert location_column == [
            spt_name,
            "Total",
            non_spt_name,
            "Total All Locations (incl. Non SPT)",
        ]

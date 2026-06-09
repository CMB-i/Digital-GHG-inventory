# NOTE: This test script requires a seeded development database to run successfully.
import os
from pathlib import Path
import sys

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from app import create_app
from app.database import db
from app.modules.RPTBLD.service import (
    create_report_template,
    get_report_template,
    update_report_template,
    delete_report_template,
    list_report_templates,
    generate_report_data,
    export_report_to_excel,
    get_missing_submissions,
    _get_user_allowed_sites
)
from app.modules.RPTBLD.model import ReportTemplate
from app.modules.USRMGMT.model import User
from app.modules.SITEMST.model import Site
from app.modules.FORMBLD.model import Form, FormVersion
from app.modules.PERIOD.model import ReportingPeriod
from app.modules.SUBMIT.model import Submission, SubmissionValue
from app.modules.ACCESS.model import AccessMatrix
from datetime import datetime

def run_tests():
    app = create_app()
    with app.app_context():
        print("Starting Phase 9 Reports & Dashboard service tests...")

        # 1. Fetch seed users
        admin_user = User.query.filter_by(email="admin@example.com").first()
        if not admin_user:
            print("FAILED: admin@example.com user not found. Please run seed script first.")
            sys.exit(1)

        print(f"Using Admin User: {admin_user.full_name} (ID: {admin_user.id})")
        # Clean up any existing test templates from previous aborted runs
        ReportTemplate.query.filter(ReportTemplate.code.in_(["test_temp_code_unique_123", "agg_temp_code_123"])).delete(synchronize_session=False)
        db.session.commit()

        # 2. Test CRUD operations for report templates
        print("\n--- Test 1: Report Template CRUD ---")
        name = "Test Report Temp"
        code = "test_temp_code_unique_123"
        desc = "A report to test Phase 9 aggregator"
        scope_type = "global"
        config = {
            "form_ids": [1],
            "site_ids": [],
            "start_year": 2026,
            "start_month": 1,
            "end_year": 2026,
            "end_month": 12
        }

        # Create
        template = create_report_template(
            name=name,
            code=code,
            description=desc,
            scope_type=scope_type,
            scope_site_id=None,
            config_json=config,
            user_id=admin_user.id
        )
        db.session.commit()
        print(f"Created template: {template.name} (ID: {template.id})")
        assert template.code == code
        assert template.config_json["start_year"] == 2026

        # Read / Get
        fetched = get_report_template(template.id)
        assert fetched is not None
        assert fetched.name == name

        # Update
        updated_name = "Updated Test Report Temp"
        config["end_month"] = 6
        updated = update_report_template(
            template_id=template.id,
            name=updated_name,
            description=fetched.description,
            scope_type=fetched.scope_type,
            scope_site_id=fetched.scope_site_id,
            config_json=config,
            user_id=admin_user.id
        )
        db.session.commit()
        print(f"Updated template: {updated.name}")
        assert updated.name == updated_name
        assert updated.config_json["end_month"] == 6

        # List
        all_templates = list_report_templates(admin_user.id)
        assert any(t.id == template.id for t in all_templates)
        print("List report templates verified successfully.")

        # 3. Test data aggregation (generate_report_data)
        print("\n--- Test 2: Data Aggregation & Approved-Only Check ---")
        # Ensure we have a site, form, and period
        site = Site.query.filter_by(is_deleted=False).first()
        form = Form.query.filter_by(is_deleted=False).first()
        period = ReportingPeriod.query.filter_by(site_id=site.id, is_deleted=False).first()

        if not site or not form or not period:
            print("WARNING: Insufficient seed data to fully execute aggregation tests. Seeding temporary data...")
            # If missing, we will seed dynamic test components
            if not site:
                site = Site(name="Test Site Alpha", code="TSA", is_deleted=False)
                db.session.add(site)
            if not form:
                form = Form(name="Test Diesel Form", code="tdf", current_version_id=1, is_deleted=False)
                db.session.add(form)
            if not period:
                period = ReportingPeriod(site_id=site.id, year=2026, month=5, status="OPEN", is_deleted=False, created_by=admin_user.id)
                db.session.add(period)
            db.session.commit()

        print(f"Using Site: {site.name} (ID: {site.id})")
        print(f"Using Form: {form.name} (ID: {form.id})")
        print(f"Using Period: {period.year}-{period.month} (ID: {period.id})")

        # Query or create two new periods to avoid colliding with seed data's active submissions
        p2_month = period.month + 1 if period.month <= 10 else period.month - 1
        p3_month = period.month + 2 if period.month <= 10 else period.month - 2
        p2_year = period.year
        p3_year = period.year

        period2 = ReportingPeriod.query.filter_by(site_id=site.id, year=p2_year, month=p2_month, is_deleted=False).first()
        if not period2:
            period2 = ReportingPeriod(site_id=site.id, year=p2_year, month=p2_month, status="OPEN", is_deleted=False, created_by=admin_user.id)
            db.session.add(period2)
            db.session.flush()

        period3 = ReportingPeriod.query.filter_by(site_id=site.id, year=p3_year, month=p3_month, is_deleted=False).first()
        if not period3:
            period3 = ReportingPeriod(site_id=site.id, year=p3_year, month=p3_month, status="OPEN", is_deleted=False, created_by=admin_user.id)
            db.session.add(period3)
            db.session.flush()

        print(f"Using Site: {site.name} (ID: {site.id})")
        print(f"Using Form: {form.name} (ID: {form.id})")
        print(f"Using Period 1 (Approved): {period3.year}-{period3.month} (ID: {period3.id})")
        print(f"Using Period 2 (Draft): {period2.year}-{period2.month} (ID: {period2.id})")

        # Setup specific template config for test
        test_config = {
            "form_ids": [form.id],
            "site_ids": [site.id],
            "start_year": min(period2.year, period3.year),
            "start_month": min(period2.month, period3.month),
            "end_year": max(period2.year, period3.year),
            "end_month": max(period2.month, period3.month)
        }

        agg_template = create_report_template(
            name="Aggregation Temp",
            code="agg_temp_code_123",
            description="Agg test",
            scope_type="global",
            scope_site_id=None,
            config_json=test_config,
            user_id=admin_user.id
        )
        db.session.commit()

        # Create user access matrix row to make sure admin has access
        # Wait, admin typically has global access, let's verify if there is an access row
        access_row = AccessMatrix.query.filter_by(user_id=admin_user.id, entity_type="report", is_deleted=False).first()
        if not access_row:
            access_row = AccessMatrix(
                user_id=admin_user.id,
                entity_type="report",
                scope_type="global",
                can_view=True,
                can_export=True,
                is_deleted=False
            )
            db.session.add(access_row)
            db.session.commit()

        # Query or create a WorkflowVersion to avoid NotNull constraint violations on Submission
        from app.modules.WFLWBLD.model import Workflow, WorkflowVersion
        wf_ver = WorkflowVersion.query.first()
        if not wf_ver:
            wf = Workflow.query.first()
            if not wf:
                wf = Workflow(name="Test Workflow", code="test_wf_123", is_deleted=False)
                db.session.add(wf)
                db.session.flush()
            wf_ver = WorkflowVersion(workflow_id=wf.id, version_number=1)
            db.session.add(wf_ver)
            db.session.flush()

        # Clean up any existing test submissions on these periods
        Submission.query.filter(Submission.site_id == site.id, Submission.form_id == form.id, Submission.reporting_period_id.in_([period2.id, period3.id])).delete(synchronize_session=False)
        db.session.commit()

        # Create two submissions for the period: one Draft and one Approved
        draft_sub = Submission(
            site_id=site.id,
            form_id=form.id,
            reporting_period_id=period2.id,
            form_version_id=form.current_version_id or 1,
            workflow_version_id=wf_ver.id,
            status="Draft",
            is_deleted=False,
            submitted_by=admin_user.id,
            created_by=admin_user.id
        )
        approved_sub = Submission(
            site_id=site.id,
            form_id=form.id,
            reporting_period_id=period3.id,
            form_version_id=form.current_version_id or 1,
            workflow_version_id=wf_ver.id,
            status="Approved",
            is_locked=True,
            is_deleted=False,
            submitted_by=admin_user.id,
            created_by=admin_user.id
        )
        db.session.add(draft_sub)
        db.session.add(approved_sub)
        db.session.commit()

        # Add a value set entry to the approved submission
        from app.modules.FORMBLD.model import FieldVersion
        fv = FieldVersion.query.filter_by(is_deleted=False).first()
        if fv:
            # Align form versions so that the aggregator can query the fields properly
            approved_sub.form_version_id = fv.form_version_id
            draft_sub.form_version_id = fv.form_version_id
            db.session.commit()
        else:
            raise Exception("No active FieldVersion found in database to execute tests.")

        val1 = SubmissionValue(
            submission_id=approved_sub.id,
            field_id=fv.field_id,
            field_version_id=fv.id,
            raw_value="1500.50",
            calculated_value="1500.50",
            created_by=admin_user.id
        )
        val2 = SubmissionValue(
            submission_id=draft_sub.id,
            field_id=fv.field_id,
            field_version_id=fv.id,
            raw_value="999.99",
            calculated_value="999.99",
            created_by=admin_user.id
        )
        db.session.add(val1)
        db.session.add(val2)
        db.session.commit()

        # Run aggregation
        data = generate_report_data(agg_template.id, admin_user.id)
        print(f"Aggregated records found: {len(data)}")
        # Check that it ONLY matches the Approved one (1500.50) and not the Draft one (999.99)
        assert len(data) > 0, "Approved submission value should have been aggregated."

        found_approved = False
        found_draft = False
        for record in data:
            if record["value"] == 1500.50:
                found_approved = True
            if record["value"] == 999.99:
                found_draft = True

        assert found_approved, "Approved submission value was not aggregated."
        assert not found_draft, "Draft submission value was incorrectly aggregated!"
        print("Aggregation approved-only check PASSED.")

        # 4. Test Excel Export
        print("\n--- Test 3: Excel Export via openpyxl ---")
        xlsx_bytes = export_report_to_excel(agg_template.id, admin_user.id)
        assert len(xlsx_bytes) > 0, "Excel output is empty."
        print(f"Generated Excel file size: {len(xlsx_bytes)} bytes")

        # Try importing openpyxl and loading the workbook from bytes
        import io
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert "Overview" in wb.sheetnames
        assert "Report Data" in wb.sheetnames
        print("Excel workbook parsed successfully and sheet names verified.")

        # 5. Test Missing Submission Tracker
        print("\n--- Test 4: Missing Submission Checker ---")
        # Ensure we have site permissions for submission module
        sub_access = AccessMatrix.query.filter_by(user_id=admin_user.id, entity_type="submission", is_deleted=False).first()
        if not sub_access:
            sub_access = AccessMatrix(
                user_id=admin_user.id,
                entity_type="submission",
                scope_type="global",
                can_view=True,
                can_submit=True,
                is_deleted=False
            )
            db.session.add(sub_access)
            db.session.commit()

        # Ensure that the form lists the site as applicable in its description
        # Form description needs to hold {"sites": [site.id]}
        import json
        form.description = json.dumps({"sites": [site.id]})
        db.session.commit()

        # Get missing submissions
        missing = get_missing_submissions(admin_user.id)
        print(f"Total missing submission entries tracked: {len(missing)}")

        # Clean up database records in correct order to avoid ForeignKeyViolations
        # Delete submission values (children) first
        db.session.delete(val1)
        db.session.delete(val2)
        db.session.commit()

        # Delete templates and submissions (parents)
        db.session.delete(template)
        db.session.delete(agg_template)
        db.session.delete(approved_sub)
        db.session.delete(draft_sub)
        db.session.commit()
        print("\nAll database cleanups completed successfully.")
        print("All Phase 9 Reports & Dashboard service tests PASSED successfully!")


if __name__ == "__main__":
    run_tests()

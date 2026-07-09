#!/usr/bin/env python3
"""Seed the Jaigarh FY25-26 GHG workbook from Final Data_Jaigarh.xlsx."""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import load_workbook

from app import create_app
from app.database import db
from app.modules.ACCESS.model import AccessMatrix
from app.modules.APPROV.model import ApprovalAction
from app.modules.FORMBLD.model import Field, FieldVersion, Form, FormSection, FormVersion
from app.modules.FRMULA.model import Formula, FormulaVersion
from app.modules.NOTIFY.model import Notification
from app.modules.PERIOD.model import ReportingPeriod
from app.modules.SITEMST.model import Site
from app.modules.SUBMIT.model import (
    Submission,
    SubmissionPackage,
    SubmissionValue,
    WorkbookFieldValue,
)
from app.modules.SUBMIT.service import autosave_submission_values, compose_annual_workbook_data
from app.modules.USRMGMT.model import User
from app.modules.WFLWBLD.model import (
    Workflow,
    WorkflowLevel,
    WorkflowLevelApprover,
    WorkflowVersion,
)
from app.modules.WKBK.model import Workbook, WorkbookForm, WorkbookSite, WorkbookSiteSubmitter

from jaigarh_workbook_config import (
    ANNUAL_EXCEL_FORM_BY_FIELD,
    ANNUAL_EXCEL_VALUES,
    FORM_DEFINITIONS,
    FORMULA_CODES,
    FORMULA_DEFINITIONS,
    FY_MONTHS,
    FY_START_YEAR,
    MONTHLY_EXCEL_MAP,
    MONTHLY_EXCEL_ROWS,
    MONTHLY_FIELD_FORM_BY_CODE,
    SITE_CODE,
    SITE_NAME,
    VERIFICATION_CHECKS,
    WORKBOOK_CODE,
    WORKBOOK_NAME,
    WORKFLOW_CODE,
    formula_tokens,
    resolve_excel_path,
)


def _null_current_versions(form_ids, formula_ids, workflow_id=None):
    for form_id in form_ids:
        form = db.session.get(Form, form_id)
        if form:
            form.current_version_id = None
    for formula_id in formula_ids:
        formula = db.session.get(Formula, formula_id)
        if formula:
            formula.current_version_id = None
    if workflow_id:
        workflow = db.session.get(Workflow, workflow_id)
        if workflow:
            workflow.current_version_id = None
    db.session.flush()


def cleanup_jaigarh_data():
    """Remove prior Jaigarh seed records (idempotent)."""
    db.session.rollback()

    workbook = Workbook.query.filter_by(code=WORKBOOK_CODE).first()
    site = Site.query.filter_by(code=SITE_CODE).first()
    workflow = Workflow.query.filter_by(code=WORKFLOW_CODE).first()
    forms = Form.query.filter(Form.code.in_([f["code"] for f in FORM_DEFINITIONS])).all()
    formulas = Formula.query.filter(Formula.code.in_(FORMULA_CODES)).all()

    form_ids = [f.id for f in forms]
    formula_ids = [f.id for f in formulas]
    workflow_id = workflow.id if workflow else None

    if workbook:
        workbook.workflow_id = None
        db.session.flush()
        WorkbookForm.query.filter_by(workbook_id=workbook.id).delete(synchronize_session=False)
        WorkbookSite.query.filter_by(workbook_id=workbook.id).delete(synchronize_session=False)
        WorkbookSiteSubmitter.query.filter_by(workbook_id=workbook.id).delete(synchronize_session=False)
        Workbook.query.filter_by(id=workbook.id).delete(synchronize_session=False)

    if site:
        WorkbookFieldValue.query.filter_by(site_id=site.id).delete(synchronize_session=False)
        subs = Submission.query.filter_by(site_id=site.id).all()
        sub_ids = [s.id for s in subs]
        if sub_ids:
            SubmissionValue.query.filter(SubmissionValue.submission_id.in_(sub_ids)).delete(synchronize_session=False)
            ApprovalAction.query.filter(ApprovalAction.submission_id.in_(sub_ids)).delete(synchronize_session=False)
            Notification.query.filter(
                Notification.entity_type == "submission",
                Notification.entity_id.in_(sub_ids),
            ).delete(synchronize_session=False)
            Submission.query.filter(Submission.id.in_(sub_ids)).delete(synchronize_session=False)
        SubmissionPackage.query.filter_by(site_id=site.id).delete(synchronize_session=False)
        ReportingPeriod.query.filter_by(site_id=site.id).delete(synchronize_session=False)
        AccessMatrix.query.filter_by(scope_site_id=site.id).delete(synchronize_session=False)

    _null_current_versions(form_ids, formula_ids, workflow_id)
    db.session.commit()

    if form_ids:
        field_ids = [
            row.id
            for row in db.session.query(Field.id).filter(Field.form_id.in_(form_ids)).all()
        ]
        if field_ids:
            FieldVersion.query.filter(FieldVersion.field_id.in_(field_ids)).delete(synchronize_session=False)
            Field.query.filter(Field.id.in_(field_ids)).delete(synchronize_session=False)
        FormSection.query.filter(FormSection.form_id.in_(form_ids)).delete(synchronize_session=False)
        FormVersion.query.filter(FormVersion.form_id.in_(form_ids)).delete(synchronize_session=False)
        Form.query.filter(Form.id.in_(form_ids)).delete(synchronize_session=False)

    if formula_ids:
        FormulaVersion.query.filter(FormulaVersion.formula_id.in_(formula_ids)).delete(synchronize_session=False)
        Formula.query.filter(Formula.id.in_(formula_ids)).delete(synchronize_session=False)

    if workflow_id:
        version_ids = [
            row.id
            for row in db.session.query(WorkflowVersion.id).filter_by(workflow_id=workflow_id).all()
        ]
        if version_ids:
            level_ids = [
                row.id
                for row in db.session.query(WorkflowLevel.id)
                .filter(WorkflowLevel.workflow_version_id.in_(version_ids))
                .all()
            ]
            if level_ids:
                WorkflowLevelApprover.query.filter(
                    WorkflowLevelApprover.workflow_level_id.in_(level_ids)
                ).delete(synchronize_session=False)
                WorkflowLevel.query.filter(WorkflowLevel.id.in_(level_ids)).delete(synchronize_session=False)
            WorkflowVersion.query.filter(WorkflowVersion.id.in_(version_ids)).delete(synchronize_session=False)
        Workflow.query.filter_by(id=workflow_id).delete(synchronize_session=False)

    if site:
        Site.query.filter_by(id=site.id).delete(synchronize_session=False)

    db.session.commit()


def create_workflow(admin_user):
    workflow = Workflow(
        name="Jaigarh Approval",
        code=WORKFLOW_CODE,
        created_by=admin_user.id,
        updated_by=admin_user.id,
    )
    db.session.add(workflow)
    db.session.flush()

    workflow_version = WorkflowVersion(
        workflow_id=workflow.id,
        version_number=1,
        published_at=datetime.now(timezone.utc),
        published_by=admin_user.id,
        created_by=admin_user.id,
    )
    db.session.add(workflow_version)
    db.session.flush()

    level = WorkflowLevel(
        workflow_version_id=workflow_version.id,
        level_number=1,
        level_name="Level 1 Approver",
        approval_mode="ANY_ONE",
        created_by=admin_user.id,
        updated_by=admin_user.id,
    )
    db.session.add(level)
    db.session.flush()

    db.session.add(
        WorkflowLevelApprover(
            workflow_level_id=level.id,
            user_id=admin_user.id,
            sequence_number=1,
            created_by=admin_user.id,
            updated_by=admin_user.id,
        )
    )
    workflow.current_version_id = workflow_version.id
    db.session.flush()
    return workflow, workflow_version


def create_formulas(admin_user):
    formula_versions = {}
    for code, definition in FORMULA_DEFINITIONS.items():
        formula = Formula(
            name=definition["name"],
            code=code,
            created_by=admin_user.id,
            updated_by=admin_user.id,
        )
        db.session.add(formula)
        db.session.flush()

        expression = definition["expression"]
        version = FormulaVersion(
            formula_id=formula.id,
            version_number=1,
            expression=expression,
            tokens=formula_tokens(expression),
            published_at=datetime.now(timezone.utc),
            published_by=admin_user.id,
            created_by=admin_user.id,
        )
        db.session.add(version)
        db.session.flush()
        formula.current_version_id = version.id
        formula_versions[code] = version
    db.session.flush()
    return formula_versions


def create_forms(site_id, admin_user, formula_versions):
    registry = {}
    for form_def in FORM_DEFINITIONS:
        form = Form(
            name=form_def["name"],
            code=form_def["code"],
            description=json.dumps({"sites": [site_id]}),
            is_deleted=False,
            created_by=admin_user.id,
        )
        db.session.add(form)
        db.session.flush()

        sections = {}
        for section_def in form_def["sections"]:
            section = FormSection(
                form_id=form.id,
                name=section_def["name"],
                code=section_def["code"],
                layout_type=section_def["layout_type"],
                display_order=section_def["display_order"],
                created_by=admin_user.id,
            )
            db.session.add(section)
            db.session.flush()
            sections[section_def["code"]] = section

        form_version = FormVersion(
            form_id=form.id,
            version_number=1,
            status="Approved",
            published_at=datetime.now(timezone.utc),
            published_by=admin_user.id,
            created_by=admin_user.id,
        )
        db.session.add(form_version)
        db.session.flush()

        field_map = {}
        for order, field_def in enumerate(form_def["fields"], start=1):
            field = Field(
                form_id=form.id,
                field_code=field_def["field_code"],
                display_order=order,
                is_deleted=False,
                created_by=admin_user.id,
            )
            db.session.add(field)
            db.session.flush()

            field_config = dict(field_def.get("field_config") or {})
            formula_code = field_def.get("formula_code")
            if formula_code:
                field_config["formula_version_id"] = formula_versions[formula_code].id

            field_version = FieldVersion(
                form_version_id=form_version.id,
                field_id=field.id,
                version_number=1,
                field_name=field_def["field_name"],
                field_type=field_def["field_type"],
                field_config=field_config,
                section_id=sections[field_def["section_code"]].id,
                frequency=field_def.get("frequency", "monthly"),
                created_by=admin_user.id,
            )
            db.session.add(field_version)
            db.session.flush()
            field_map[field_def["field_code"]] = {
                "field": field,
                "field_version": field_version,
            }

        form.current_version_id = form_version.id
        registry[form_def["code"]] = {
            "form": form,
            "form_version": form_version,
            "fields": field_map,
            "display_order": form_def["display_order"],
            "sheet_label": form_def["sheet_label"],
        }
    db.session.flush()
    return registry


def create_site_and_periods(admin_user):
    site = Site(
        name=SITE_NAME,
        code=SITE_CODE,
        company_name="JSW Ports",
        description="Jaigarh port GHG inventory site",
        is_deleted=False,
        created_by=admin_user.id,
    )
    db.session.add(site)
    db.session.flush()

    for year, month in FY_MONTHS:
        db.session.add(
            ReportingPeriod(
                site_id=site.id,
                year=year,
                month=month,
                status="OPEN",
                is_deleted=False,
                created_by=admin_user.id,
            )
        )

    db.session.add(
        AccessMatrix(
            user_id=admin_user.id,
            scope_type="site",
            scope_site_id=site.id,
            entity_type="submission",
            can_create=True,
            can_submit=True,
            can_view=True,
            can_edit=True,
            can_approve=True,
            created_by=admin_user.id,
        )
    )
    db.session.flush()
    return site


def create_workbook(site, admin_user, workflow, form_registry):
    workbook = Workbook(
        name=WORKBOOK_NAME,
        code=WORKBOOK_CODE,
        status="published",
        description="Jaigarh FY25-26 GHG workbook seeded from Excel",
        workflow_id=workflow.id,
        created_by=admin_user.id,
    )
    db.session.add(workbook)
    db.session.flush()

    db.session.add(
        WorkbookSite(workbook_id=workbook.id, site_id=site.id, created_by=admin_user.id)
    )
    db.session.add(
        WorkbookSiteSubmitter(
            workbook_id=workbook.id,
            site_id=site.id,
            user_id=admin_user.id,
            created_by=admin_user.id,
        )
    )

    for form_code, entry in sorted(form_registry.items(), key=lambda item: item[1]["display_order"]):
        db.session.add(
            WorkbookForm(
                workbook_id=workbook.id,
                form_id=entry["form"].id,
                display_order=entry["display_order"],
                sheet_label=entry["sheet_label"],
            )
        )
    db.session.flush()
    return workbook


def load_excel_data(excel_path, site, admin_user, workflow_version, form_registry):
    workbook = load_workbook(excel_path, data_only=True)
    periods = {
        (period.year, period.month): period
        for period in ReportingPeriod.query.filter_by(site_id=site.id, is_deleted=False).all()
    }

    monthly_forms = {
        form_code: form_registry[form_code]
        for form_code in ("form_jai_electricity", "form_jai_fuel", "form_jai_cargo")
    }

    for field_code, (sheet_name, column, row_key) in MONTHLY_EXCEL_MAP.items():
        form_code = MONTHLY_FIELD_FORM_BY_CODE[field_code]
        entry = monthly_forms[form_code]
        field_info = entry["fields"][field_code]
        excel_rows = MONTHLY_EXCEL_ROWS[row_key]

        for row_index, (year, month) in zip(excel_rows, FY_MONTHS):
            period = periods[(year, month)]
            value = workbook[sheet_name][f"{column}{row_index}"].value
            if value is None:
                continue

            submission = Submission.query.filter_by(
                site_id=site.id,
                form_id=entry["form"].id,
                reporting_period_id=period.id,
                is_deleted=False,
            ).first()
            if not submission:
                submission = Submission(
                    site_id=site.id,
                    form_id=entry["form"].id,
                    form_version_id=entry["form_version"].id,
                    reporting_period_id=period.id,
                    workflow_version_id=workflow_version.id,
                    status="Draft",
                    is_deleted=False,
                    created_by=admin_user.id,
                )
                db.session.add(submission)
                db.session.flush()

            existing_value = SubmissionValue.query.filter_by(
                submission_id=submission.id,
                field_id=field_info["field"].id,
            ).first()
            if existing_value:
                existing_value.raw_value = str(value)
                existing_value.field_version_id = field_info["field_version"].id
                existing_value.cell_state = "draft_filled"
                existing_value.updated_by = admin_user.id
            else:
                db.session.add(
                    SubmissionValue(
                        submission_id=submission.id,
                        field_id=field_info["field"].id,
                        field_version_id=field_info["field_version"].id,
                        raw_value=str(value),
                        cell_state="draft_filled",
                        created_by=admin_user.id,
                    )
                )

    for field_code, (sheet_name, cell) in ANNUAL_EXCEL_VALUES.items():
        value = workbook[sheet_name][cell].value
        if value is None:
            continue
        form_code = ANNUAL_EXCEL_FORM_BY_FIELD[field_code]
        annual_form = form_registry[form_code]
        field_info = annual_form["fields"][field_code]
        db.session.add(
            WorkbookFieldValue(
                site_id=site.id,
                form_id=annual_form["form"].id,
                field_id=field_info["field"].id,
                field_version_id=field_info["field_version"].id,
                fy_start_year=FY_START_YEAR,
                value_text=str(value),
                numeric_value=Decimal(str(value)),
                cell_state="draft_filled",
                created_by=admin_user.id,
            )
        )

    db.session.flush()

    for form_code in ("form_jai_electricity", "form_jai_fuel", "form_jai_cargo"):
        entry = form_registry[form_code]
        submissions = Submission.query.filter_by(
            site_id=site.id,
            form_id=entry["form"].id,
            is_deleted=False,
        ).all()
        for submission in submissions:
            autosave_submission_values(submission.id, {}, admin_user.id)

    db.session.commit()


def verify_workbook(site, workbook, admin_user, excel_path):
    excel_wb = load_workbook(excel_path, data_only=True)
    failures = []
    passes = []

    for check in VERIFICATION_CHECKS:
        form_code = check["form_code"]
        form = Form.query.filter_by(code=form_code).first()
        if not form:
            failures.append(f"{check['field_code']}: form {form_code} not found")
            continue

        expected = excel_wb[check["excel_sheet"]][check["excel_cell"]].value

        if check.get("aggregation") == "sum_months":
            payload = compose_annual_workbook_data(
                admin_user.id,
                site.id,
                workbook.id,
                FY_START_YEAR,
                selected_form_id=form.id,
            )
            field_code = check["field_code"]
            monthly_sum = 0.0
            found_values = 0
            for row in payload.get("rows") or []:
                cell = (row.get("values") or {}).get(field_code)
                if cell is None:
                    continue
                raw = cell.get("value") if isinstance(cell, dict) else cell
                if raw in (None, ""):
                    continue
                monthly_sum += float(raw)
                found_values += 1
            if found_values == 0:
                failures.append(
                    f"{check['field_code']}: no monthly values to sum (expected {expected})"
                )
                continue
            actual = monthly_sum
        else:
            payload = compose_annual_workbook_data(
                admin_user.id,
                site.id,
                workbook.id,
                FY_START_YEAR,
                selected_form_id=form.id,
            )
            sheet_results = {
                item["field_code"]: item["value"]
                for item in payload.get("sheet_results") or []
            }
            actual = sheet_results.get(check["field_code"])

            if actual is None:
                failures.append(
                    f"{check['field_code']}: no calculated value (expected {expected})"
                )
                continue

        actual_f = float(actual)
        expected_f = float(expected)
        delta = abs(actual_f - expected_f)
        if delta <= check["tolerance"]:
            passes.append(
                f"{check['field_code']}: {actual_f:.4f} ~= {expected_f:.4f} (delta {delta:.4f})"
            )
        else:
            failures.append(
                f"{check['field_code']}: {actual_f:.4f} != {expected_f:.4f} "
                f"(delta {delta:.4f}, tolerance {check['tolerance']})"
            )

    return passes, failures


def seed_jaigarh_workbook(excel_path, verify_only=False):
    excel_path = resolve_excel_path(excel_path)
    if not excel_path.exists():
        print(f"FAILED: Excel file not found at {excel_path}")
        sys.exit(1)

    app = create_app()
    with app.app_context():
        admin_user = User.query.filter_by(email="admin@example.com").first()
        if not admin_user:
            print("FAILED: Seed admin user not found. Run scripts/seed.py first.")
            sys.exit(1)

        if verify_only:
            site = Site.query.filter_by(code=SITE_CODE).first()
            workbook = Workbook.query.filter_by(code=WORKBOOK_CODE).first()
            if not site or not workbook:
                print("FAILED: Jaigarh workbook not seeded yet. Run without --verify-only first.")
                sys.exit(1)
            passes, failures = verify_workbook(site, workbook, admin_user, excel_path)
            for line in passes:
                print(f"PASS: {line}")
            for line in failures:
                print(f"FAIL: {line}")
            if failures:
                sys.exit(1)
            print("All verification checks passed.")
            return

        print("Cleaning up prior Jaigarh seed data...")
        cleanup_jaigarh_data()

        print("Creating Jaigarh site and reporting periods...")
        site = create_site_and_periods(admin_user)

        print("Creating approval workflow...")
        workflow, workflow_version = create_workflow(admin_user)

        print("Creating formulas...")
        formula_versions = create_formulas(admin_user)

        print("Creating forms and fields...")
        form_registry = create_forms(site.id, admin_user, formula_versions)

        print("Creating workbook...")
        workbook = create_workbook(site, admin_user, workflow, form_registry)

        print(f"Loading monthly and annual data from {excel_path}...")
        load_excel_data(excel_path, site, admin_user, workflow_version, form_registry)

        print("Running verification checks...")
        passes, failures = verify_workbook(site, workbook, admin_user, excel_path)
        for line in passes:
            print(f"PASS: {line}")
        for line in failures:
            print(f"FAIL: {line}")

        if failures:
            print("Seed completed with verification failures.")
            sys.exit(1)

        print("Jaigarh workbook seeded successfully.")
        print(f"  Site: {SITE_CODE} (id={site.id})")
        print(f"  Workbook: {WORKBOOK_CODE} (id={workbook.id})")
        print(f"  FY: {FY_START_YEAR} (Apr 2025 – Mar 2026)")
        print("  Login as admin@example.com and open /module/SUBMIT/annual")


def main():
    parser = argparse.ArgumentParser(description="Seed Jaigarh FY25-26 GHG workbook from Excel.")
    parser.add_argument(
        "--excel-path",
        help="Path to Final Data_Jaigarh.xlsx (default: Downloads path)",
    )
    parser.add_argument(
        "--verify-only",
        action="store_true",
        help="Only run verification checks against an existing seed",
    )
    args = parser.parse_args()
    seed_jaigarh_workbook(args.excel_path, verify_only=args.verify_only)


if __name__ == "__main__":
    main()

import io
from datetime import datetime, timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import tuple_

from app.database import db
from app.modules.RPTBLD.model import ReportTemplate, AppConfig
from app.modules.WKBK.model import Workbook, WorkbookForm, WorkbookSite
from app.common.permissions import has_permission
from app.modules.ACCESS.service import get_user_permissions
from app.modules.SITEMST.model import Site
from app.modules.PERIOD.model import ReportingPeriod
from app.modules.FORMBLD.model import Form, Field, FieldVersion
from app.modules.SUBMIT.model import Submission, SubmissionValue, ProofDocument
from app.modules.SUBMIT.service import human_sheet_label
from app.modules.USRMGMT.model import User
from app.modules.FRMULA.model import Formula, FormulaVersion
from app.modules.FRMULA.service import (
    REPORT_CONTEXT_METRIC_KEYS,
    evaluate_formula,
    FormulaValidationError,
)

def list_report_templates(user_id):
    """
    List report templates. Scoped by user access.
    """
    # Verify user has access to view reports. Uses get_user_permissions() (rather
    # than a hand-rolled AccessMatrix scan) so a blanket entity_type == "all" grant
    # is correctly honored, same as any other entity type.
    global_perms = get_user_permissions(user_id=user_id, scope_type="global", entity_type="report")
    is_global = bool(global_perms["can_view"] or global_perms["can_export"])

    if is_global:
        return ReportTemplate.query.filter_by(is_deleted=False).order_by(ReportTemplate.id.desc()).all()

    allowed_site_ids = set()
    for site in Site.query.filter_by(is_deleted=False).all():
        perms = get_user_permissions(user_id=user_id, scope_type="site", scope_site_id=site.id, entity_type="report")
        if perms["can_view"] or perms["can_export"]:
            allowed_site_ids.add(site.id)

    # Filter templates that match the user's allowed sites
    all_templates = ReportTemplate.query.filter_by(is_deleted=False).all()
    filtered = []
    for t in all_templates:
        # If template is global, or site matches user's allowed list
        if not t.scope_site_id or t.scope_site_id in allowed_site_ids:
            filtered.append(t)
    return filtered

def get_report_template(template_id):
    return ReportTemplate.query.filter_by(id=template_id, is_deleted=False).one_or_none()


def validate_row_groups(row_groups):
    """Exactly one group in a non-empty row_groups list must be the
    reference base. An empty/absent row_groups (templates not yet using
    this feature) is valid and skips the check entirely."""
    if not row_groups:
        return
    reference_base_count = sum(1 for g in row_groups if g.get("is_reference_base"))
    if reference_base_count != 1:
        raise ValueError(
            f"Exactly one row group must have is_reference_base=true (found {reference_base_count})."
        )


def validate_computed_columns(computed_columns):
    """Every computed column's formula_id must point at a live, published
    context='report' Formula -- a context='field' formula_id here is a
    config error, not silently ignored.

    kind="cross_site" is the one exception: these columns (% Contribution of
    Total Emissions, Variation from Average Intensity) are genuine cross-row
    aggregations against a specific row_group's own total -- a per-row
    Formula (evaluated one row at a time against that row's own metrics plus
    group subtotals) cannot express "always divide by the SPT group's total,
    even for Non-SPT rows" or "only populate this for rows in the reference
    group." compose_cross_site_intensity_report computes their values
    directly in Python instead; they carry no formula_id at all, by design."""
    if not computed_columns:
        return
    for col in computed_columns:
        col_id = col.get("id")
        if not col_id:
            raise ValueError("Each computed column requires an 'id'.")
        if col.get("kind") == "cross_site":
            continue
        formula_id = col.get("formula_id")
        if formula_id is None:
            raise ValueError(f"Computed column '{col_id}' requires a 'formula_id'.")
        formula = Formula.query.filter_by(id=formula_id, is_deleted=False).one_or_none()
        if not formula:
            raise ValueError(f"Computed column '{col_id}' references formula_id {formula_id}, which does not exist.")
        if formula.context != "report":
            raise ValueError(
                f"Computed column '{col_id}' references formula_id {formula_id}, "
                f"which has context='{formula.context}', not 'report'."
            )


def validate_metric_aliases(config_json):
    """Every metric_aliases key must be one of the canonical report metric
    keys or a computed_columns 'id' (the PNP/MELT-style per-site override
    case). Every entry's site_id must appear in some row_groups entry's
    site_ids (an alias for a site outside every row group is orphaned
    config) and every field_id must resolve to a live, non-deleted Field.
    op='sum' requires >=2 field_ids; op='single' requires exactly 1."""
    config_json = config_json or {}
    metric_aliases = config_json.get("metric_aliases")
    if not metric_aliases:
        return

    valid_site_ids = set()
    for group in config_json.get("row_groups") or []:
        valid_site_ids.update(group.get("site_ids") or [])

    computed_column_ids = {c.get("id") for c in (config_json.get("computed_columns") or [])}
    valid_metric_keys = REPORT_CONTEXT_METRIC_KEYS | computed_column_ids

    for metric_key, entries in metric_aliases.items():
        if metric_key not in valid_metric_keys:
            raise ValueError(
                f"metric_aliases key '{metric_key}' is not a canonical report metric or a defined computed column id."
            )
        for entry in entries or []:
            site_id = entry.get("site_id")
            if site_id not in valid_site_ids:
                raise ValueError(
                    f"metric_aliases entry for '{metric_key}' references site_id {site_id}, "
                    "which is not present in any row group."
                )
            field_ids = entry.get("field_ids") or []
            op = entry.get("op")
            if op == "single":
                if len(field_ids) != 1:
                    raise ValueError(
                        f"metric_aliases entry for '{metric_key}'/site {site_id}: op='single' requires exactly 1 field_id."
                    )
            elif op == "sum":
                if len(field_ids) < 2:
                    raise ValueError(
                        f"metric_aliases entry for '{metric_key}'/site {site_id}: op='sum' requires at least 2 field_ids."
                    )
            else:
                raise ValueError(
                    f"metric_aliases entry for '{metric_key}'/site {site_id}: op must be 'single' or 'sum', got {op!r}."
                )
            for field_id in field_ids:
                field = Field.query.filter_by(id=field_id, is_deleted=False).one_or_none()
                if not field:
                    raise ValueError(
                        f"metric_aliases entry for '{metric_key}'/site {site_id} references field_id {field_id}, "
                        "which does not exist or is deleted."
                    )


def _apply_metric_alias_defaults(config_json):
    """verified defaults to false on creation -- an admin has to explicitly
    confirm each mapping. Mutates config_json in place."""
    metric_aliases = (config_json or {}).get("metric_aliases")
    if not metric_aliases:
        return
    for entries in metric_aliases.values():
        for entry in entries or []:
            entry.setdefault("verified", False)


def validate_report_config(config_json):
    """Single validation entry point for everything config_json can hold --
    called from both create_report_template and update_report_template so
    there is exactly one place writing (and therefore validating) this
    config, per the app's 'one validated backend' rule."""
    config_json = config_json or {}
    validate_row_groups(config_json.get("row_groups") or [])
    validate_computed_columns(config_json.get("computed_columns") or [])
    validate_metric_aliases(config_json)
    return config_json


def get_emission_factor_version():
    """
    Display/documentation only -- the source workbook's emission-factor
    cells are reference footnotes, not active formula inputs (Scope-1/2
    values arrive already computed from upstream). Never wired into
    evaluate_formula's names dict.
    """
    row = AppConfig.query.filter_by(config_key="cea_emission_factor_version").first()
    return row.config_value if row else "V21"  # documented default, not a silent guess


def create_report_template(name, code, description, scope_type, scope_site_id, config_json, user_id):
    if not name or not name.strip():
        raise ValueError("Report name is required.")
    if not code or not code.strip():
        raise ValueError("Report code is required.")

    existing = ReportTemplate.query.filter_by(code=code, is_deleted=False).first()
    if existing:
        raise ValueError(f"Report template with code '{code}' already exists.")

    # Validation config
    if not config_json:
        config_json = {}
    _apply_metric_alias_defaults(config_json)
    validate_report_config(config_json)

    template = ReportTemplate(
        name=name.strip(),
        code=code.strip(),
        description=description,
        scope_type=scope_type or "global",
        scope_site_id=scope_site_id,
        config_json=config_json,
        created_by=user_id,
        updated_by=user_id
    )
    db.session.add(template)
    db.session.flush()
    return template

def update_report_template(template_id, name, description, scope_type, scope_site_id, config_json, user_id):
    t = get_report_template(template_id)
    if not t:
        raise ValueError("Report template not found.")

    if name:
        t.name = name.strip()
    # Mirrors name's own guard above -- an empty/blank description from the
    # caller leaves the existing one alone rather than silently wiping it.
    # This was a real regression: a save where the description textarea
    # happened to be empty at that moment blanked out a previously-set
    # description with no warning.
    if description and description.strip():
        t.description = description.strip()
    t.scope_type = scope_type or "global"
    t.scope_site_id = scope_site_id
    if config_json is not None:
        _apply_metric_alias_defaults(config_json)
        validate_report_config(config_json)
        t.config_json = config_json
    t.updated_by = user_id
    db.session.flush()
    return t

def delete_report_template(template_id, user_id):
    t = get_report_template(template_id)
    if not t:
        raise ValueError("Report template not found.")
    t.is_deleted = True
    t.deleted_by = user_id
    t.deleted_at = datetime.now(timezone.utc)
    t.delete_reason = "Deleted by user"
    return True

def _get_user_allowed_sites(user_id, entity_type="report"):
    # Uses get_user_permissions() (rather than a hand-rolled AccessMatrix scan) so
    # a blanket entity_type == "all" grant is correctly honored, same as any other
    # entity type.
    global_perms = get_user_permissions(user_id=user_id, scope_type="global", entity_type=entity_type)
    if global_perms["can_view"] or global_perms["can_export"] or global_perms["can_submit"] or global_perms["can_create"]:
        active_sites = Site.query.filter_by(is_deleted=False).all()
        return {s.id for s in active_sites}, True

    allowed_site_ids = set()
    for site in Site.query.filter_by(is_deleted=False).all():
        perms = get_user_permissions(user_id=user_id, scope_type="site", scope_site_id=site.id, entity_type=entity_type)
        if perms["can_view"] or perms["can_export"] or perms["can_submit"] or perms["can_create"]:
            allowed_site_ids.add(site.id)
    return allowed_site_ids, False

def generate_report_data(template_id, user_id):
    """
    Gathers approved data for the given template. Scopes by user permissions.
    """
    t = get_report_template(template_id)
    if not t:
        raise ValueError("Report template not found.")

    allowed_site_ids, is_global = _get_user_allowed_sites(user_id, "report")

    config = t.config_json or {}
    form_ids = config.get("form_ids", [])
    site_ids = config.get("site_ids", [])
    start_year = config.get("start_year")
    start_month = config.get("start_month")
    end_year = config.get("end_year")
    end_month = config.get("end_month")

    # Filter sites by user's permitted sites
    if site_ids:
        query_site_ids = [sid for sid in site_ids if sid in allowed_site_ids]
    else:
        query_site_ids = list(allowed_site_ids)

    if not query_site_ids:
        return []

    # Query approved submissions
    sub_query = Submission.query.filter(
        Submission.site_id.in_(query_site_ids),
        Submission.status == "Approved",
        Submission.is_locked == True,
        Submission.is_deleted == False
    )

    if form_ids:
        sub_query = sub_query.filter(Submission.form_id.in_(form_ids))

    submissions = sub_query.all()

    # Filter by date range (joining reporting period)
    filtered_subs = []
    for sub in submissions:
        p = ReportingPeriod.query.get(sub.reporting_period_id)
        if not p or p.is_deleted:
            continue

        p_val = p.year * 12 + p.month
        if start_year and start_month:
            if p_val < (start_year * 12 + start_month):
                continue
        if end_year and end_month:
            if p_val > (end_year * 12 + end_month):
                continue
        filtered_subs.append((sub, p))

    # Gather values
    results = []
    from app.modules.SUBMIT.service import format_period_label

    sites_cache = {s.id: s for s in Site.query.filter_by(is_deleted=False).all()}
    forms_cache = {f.id: f for f in Form.query.filter_by(is_deleted=False).all()}

    for sub, p in filtered_subs:
        site = sites_cache.get(sub.site_id)
        form = forms_cache.get(sub.form_id)
        if not site or not form:
            continue

        # Get field configurations
        fields = (
            FieldVersion.query.with_entities(FieldVersion, Field)
            .join(Field, Field.id == FieldVersion.field_id)
            .filter(
                FieldVersion.form_version_id == sub.form_version_id,
                FieldVersion.is_deleted == False,
                Field.is_deleted == False
            )
            .all()
        )

        # Load values
        vals = SubmissionValue.query.filter_by(submission_id=sub.id).all()
        vals_map = {v.field_id: v for v in vals}

        for fv, f in fields:
            val_obj = vals_map.get(f.id)
            if not val_obj:
                continue

            unit = fv.field_config.get("unit") or "—"

            # Decide display value
            if fv.field_type == "calculated":
                display_val = float(val_obj.calculated_value) if val_obj.calculated_value is not None else None
            elif fv.field_type == "file":
                proof = ProofDocument.query.filter_by(submission_id=sub.id, field_id=f.id, is_deleted=False).first()
                display_val = proof.original_name if proof else "No Upload"
            else:
                try:
                    display_val = float(val_obj.raw_value) if val_obj.raw_value is not None else None
                except ValueError:
                    display_val = val_obj.raw_value

            results.append({
                "period_label": format_period_label(p.year, p.month),
                "site_id": site.id,
                "site_name": site.name,
                "form_name": human_sheet_label(form),
                "field_id": f.id,
                "field_code": f.field_code,
                "field_name": fv.field_name,
                "field_type": fv.field_type,
                "value": display_val,
                "unit": unit
            })

    # Sort by Period, Site, Form, Display Order
    results.sort(key=lambda x: (x["period_label"], x["site_name"], x["form_name"], x["field_name"]))
    return results


def _find_site_alias_entry(entries, site_id):
    for entry in entries or []:
        if entry.get("site_id") == site_id:
            return entry
    return None


def _resolve_alias_entry_value(entry, flat_index, site_id):
    """Resolves one metric_aliases entry against the flat-row index for a
    given site. Missing/non-numeric source values are excluded rather than
    treated as 0 (app-wide rule) -- a resolved value of None means every
    constituent field_id was blank, not that they summed to zero. Returns
    (value_or_None, verified_bool)."""
    if entry is None:
        return None, False

    numeric_values = []
    for field_id in entry.get("field_ids") or []:
        raw = flat_index.get((site_id, field_id))
        if isinstance(raw, (int, float)):
            numeric_values.append(raw)

    if not numeric_values:
        value = None
    elif entry.get("op") == "sum":
        value = sum(numeric_values)
    else:
        value = numeric_values[0]

    return value, bool(entry.get("verified", False))


def _formula_published_expression_and_tokens(formula_id):
    """
    Returns (expression, required_token_names) for a formula's published
    version, or (None, None) if it has none. required_token_names comes
    straight from FormulaVersion.tokens -- the same dict publish_formula_version
    itself validates report-context tokens against (see its tokens_keys =
    set((version.tokens or {}).keys()) check) -- not a freshly-written
    expression scanner.
    """
    formula = Formula.query.filter_by(id=formula_id, is_deleted=False).one_or_none()
    if not formula or not formula.current_version_id:
        return None, None
    version = FormulaVersion.query.get(formula.current_version_id)
    if not version:
        return None, None
    return version.expression, set((version.tokens or {}).keys())


def _evaluate_computed_columns(computed_columns, metric_aliases, flat_index, site_id, own_metrics, group_subtotal_names):
    """
    Evaluates every computed_columns entry for one row (a per-site row when
    site_id is not None, or a group-subtotal/grand-total row when it is).

    own_metrics: {metric_key: value_or_None} for this row alone.
    group_subtotal_names: {"{group_id}__{metric_key}": value} across every
    group's subtotal -- available to every row's formula regardless of which
    group the row belongs to.

    The PNP/MELT override case: when site_id is given and metric_aliases has
    an entry keyed by this computed column's own 'id' for this site, that
    entry's resolved value is used verbatim instead of evaluating the
    formula -- marked "source": "override" so this is visible at every layer
    above this function, not a silent substitution. Override never applies
    to group-subtotal/grand-total rows (there is no single sourced cell to
    substitute for an aggregate).

    A formula error on one cell does not abort the report -- it's caught
    per-cell and returned as an explicit {"value": None, "error": "..."},
    distinguishable from a merely-missing alias (which has "error": None).

    A genuinely-absent metric (a site with no petrol/IFO consumption at all,
    say) is expected, ordinary shape -- not a formula error. Before
    evaluating, any of the formula's own required tokens (FormulaVersion.tokens,
    the same set publish_formula_version validates against) that aren't
    present in `names` means this cell has nothing to compute from, so it
    degrades to a blank {"value": None, "error": None} cell -- the same
    "missing is not an error" contract kind="cross_site" columns already give
    N/P -- rather than reaching evaluate_formula and raising
    "Unknown formula variable".
    """
    names = {k: v for k, v in own_metrics.items() if v is not None}
    names.update(group_subtotal_names)

    result = {}
    for col in computed_columns:
        if col.get("kind") == "cross_site":
            continue
        col_id = col["id"]
        formula_id = col["formula_id"]

        if site_id is not None:
            override_entry = _find_site_alias_entry((metric_aliases or {}).get(col_id), site_id)
            if override_entry is not None:
                value, verified = _resolve_alias_entry_value(override_entry, flat_index, site_id)
                result[col_id] = {"value": value, "source": "override", "verified": verified, "error": None}
                continue

        expression, required_tokens = _formula_published_expression_and_tokens(formula_id)
        if expression is None:
            result[col_id] = {
                "value": None, "source": "computed", "verified": None,
                "error": f"Formula {formula_id} has no published version.",
            }
            continue

        if required_tokens - names.keys():
            result[col_id] = {"value": None, "source": "computed", "verified": None, "error": None}
            continue

        try:
            value = evaluate_formula(expression, names)
            result[col_id] = {"value": value, "source": "computed", "verified": None, "error": None}
        except FormulaValidationError as error:
            result[col_id] = {"value": None, "source": "computed", "verified": None, "error": str(error)}

    return result


def pivot_report_data(template_id, user_id):
    """
    Builds a row-grouped, metric-aliased, computed-column pivot on top of
    generate_report_data()'s flat rows (reused, not re-queried) plus
    config_json's row_groups / metric_aliases / computed_columns. Does not
    replace or modify generate_report_data -- its flat-row contract is
    unchanged for existing consumers.

    Returns:
    {
      "template_id": int,
      "row_groups": [
        {
          "id": str, "label": str, "is_reference_base": bool,
          "include_in_grand_total": bool,
          "site_rows": [
            {
              "site_id": int, "site_name": str,
              "metrics": {
                  metric_key: {"value": float | None, "verified": bool}, ...
              },
              "computed": {
                  col_id: {
                      "value": float | None,
                      "source": "computed" | "override",
                      "verified": bool | None,   # only meaningful for "override"
                      "error": str | None,
                  }, ...
              },
            }, ...
          ],
          "subtotal": {
            "label": str,
            "metrics": {metric_key: float | None, ...},   # plain values, not per-cell dicts --
                                                            # an aggregate has no single "verified" flag
            "computed": {col_id: {...same shape as above, "source" always "computed"...}, ...},
          },
        }, ...
      ],
      "grand_total": {
        "metrics": {metric_key: float | None, ...},
        "computed": {col_id: {...}, ...},
      },
    }

    A metric/cell is None whenever every value feeding it was missing --
    never coerced to 0, at any level (site cell, group subtotal, grand
    total), per the app-wide "missing values are never treated as zero" rule.
    """
    flat_rows = generate_report_data(template_id, user_id)

    t = get_report_template(template_id)
    config = t.config_json or {}
    row_groups_config = config.get("row_groups") or []
    metric_aliases = config.get("metric_aliases") or {}
    computed_columns = config.get("computed_columns") or []

    flat_index = {}
    for row in flat_rows:
        site_id = row.get("site_id")
        field_id = row.get("field_id")
        if site_id is None or field_id is None:
            continue
        flat_index[(site_id, field_id)] = row.get("value")

    all_site_ids = {sid for g in row_groups_config for sid in (g.get("site_ids") or [])}
    site_names = {s.id: s.name for s in Site.query.filter(Site.id.in_(all_site_ids or [0])).all()}

    group_subtotal_metrics = {}
    result_groups = []

    for group in row_groups_config:
        group_id = group["id"]
        site_rows = []
        for site_id in group.get("site_ids") or []:
            own_metrics = {}
            metrics_out = {}
            for metric_key in REPORT_CONTEXT_METRIC_KEYS:
                entry = _find_site_alias_entry(metric_aliases.get(metric_key), site_id)
                value, verified = _resolve_alias_entry_value(entry, flat_index, site_id)
                own_metrics[metric_key] = value
                metrics_out[metric_key] = {"value": value, "verified": verified}

            site_rows.append({
                "site_id": site_id,
                "site_name": site_names.get(site_id),
                "metrics": metrics_out,
                "_own_metrics": own_metrics,
            })

        subtotal_metrics = {}
        for metric_key in REPORT_CONTEXT_METRIC_KEYS:
            values = [
                r["metrics"][metric_key]["value"]
                for r in site_rows
                if r["metrics"][metric_key]["value"] is not None
            ]
            subtotal_metrics[metric_key] = sum(values) if values else None

        group_subtotal_metrics[group_id] = subtotal_metrics
        result_groups.append({
            "id": group_id,
            "label": group.get("label"),
            "subtotal_label": group.get("subtotal_label"),
            "include_in_grand_total": bool(group.get("include_in_grand_total")),
            "is_reference_base": bool(group.get("is_reference_base")),
            "site_rows": site_rows,
        })

    # Available to every row's formula regardless of which group the row belongs to.
    group_subtotal_names = {
        f"{group_id}__{metric_key}": value
        for group_id, metrics in group_subtotal_metrics.items()
        for metric_key, value in metrics.items()
        if value is not None
    }

    for group_result in result_groups:
        for site_row in group_result["site_rows"]:
            own_metrics = site_row.pop("_own_metrics")
            site_row["computed"] = _evaluate_computed_columns(
                computed_columns, metric_aliases, flat_index,
                site_row["site_id"], own_metrics, group_subtotal_names,
            )

        group_result["subtotal"] = {
            "label": group_result.pop("subtotal_label"),
            "metrics": group_subtotal_metrics[group_result["id"]],
            "computed": _evaluate_computed_columns(
                computed_columns, metric_aliases, flat_index,
                None, group_subtotal_metrics[group_result["id"]], group_subtotal_names,
            ),
        }

    included_groups = [g for g in result_groups if g["include_in_grand_total"]]
    grand_total_metrics = {}
    for metric_key in REPORT_CONTEXT_METRIC_KEYS:
        values = [
            group_subtotal_metrics[g["id"]][metric_key]
            for g in included_groups
            if group_subtotal_metrics[g["id"]][metric_key] is not None
        ]
        grand_total_metrics[metric_key] = sum(values) if values else None

    grand_total_computed = _evaluate_computed_columns(
        computed_columns, metric_aliases, flat_index, None, grand_total_metrics, group_subtotal_names,
    )

    return {
        "template_id": template_id,
        "row_groups": result_groups,
        "grand_total": {
            "metrics": grand_total_metrics,
            "computed": grand_total_computed,
        },
    }


def _current_calendar_fy_start_year():
    today = datetime.now(timezone.utc).date()
    return today.year if today.month >= 4 else today.year - 1


def latest_fy_start_year_with_data(site_ids, lookback_years=6):
    """
    Picks a sensible default fy_start_year for the cross-site preview: the
    calendar-derived "current FY" (April-March) is NOT a safe default,
    because a brand-new FY commonly has zero submissions for weeks/months
    after it starts while the just-closed FY still holds all the real data
    (verified in production: today's calendar FY had one stray row while the
    entire prior FY was fully populated across all 10 sites -- defaulting to
    "today's FY" silently produced an all-blank report). Instead, walk
    backwards from the current calendar FY and return the first one that
    actually has at least one Submission for any of these sites; falls back
    to the current calendar FY if none of the lookback years have any data
    at all (a genuinely new template with nothing submitted yet).
    """
    from app.modules.SUBMIT import service as submit_svc

    if not site_ids:
        return _current_calendar_fy_start_year()

    current_fy = _current_calendar_fy_start_year()
    for fy_start_year in range(current_fy, current_fy - lookback_years, -1):
        month_keys = [(item["year"], item["month"]) for item in submit_svc._fy_months(fy_start_year)]
        exists = (
            db.session.query(Submission.id)
            .join(ReportingPeriod, ReportingPeriod.id == Submission.reporting_period_id)
            .filter(
                Submission.site_id.in_(site_ids),
                Submission.is_deleted == False,
                ReportingPeriod.is_deleted == False,
                tuple_(ReportingPeriod.year, ReportingPeriod.month).in_(month_keys),
            )
            .first()
        )
        if exists:
            return fy_start_year

    return current_fy


def _site_active_workbook_id(site_id):
    """
    Resolves the site's currently-assigned workbook via Workbook.is_active --
    deliberately not Workbook.status (live/draft for data-entry purposes),
    which this read-only report must not gate on. A site can have more than
    one WorkbookSite row (an old/superseded assignment plus the real current
    one); is_active is what distinguishes them.
    """
    workbook_site = (
        WorkbookSite.query.join(Workbook, Workbook.id == WorkbookSite.workbook_id)
        .filter(WorkbookSite.site_id == site_id, Workbook.is_active == True)
        .first()
    )
    return workbook_site.workbook_id if workbook_site else None


def _site_flat_index_for_fy(site_id, fy_start_year, include_unapproved):
    """
    Builds a {field_id: value} map for one site's entire workbook, correctly
    aggregated across the whole FY via SUM_MONTHS -- this is the fix for
    generate_report_data/pivot_report_data's flat_index, which is last-write-
    wins across a multi-month date range (each (site_id, field_id) key gets
    silently overwritten by whichever submission is processed last, instead
    of being summed), so it would produce a single month's number instead of
    a cumulative FY total. Bypassed entirely here, not patched: this reads
    each sheet's real, already-tested per-site formula engine
    (_field_payload/_compose_sheet_results/synthesize_automatic_fy_totals,
    the same functions compose_annual_workbook_data uses) directly, which is
    SUM_MONTHS-correct by construction and has zero dependency on
    Workbook.status (verified: none of these functions reference it).

    include_unapproved=False restricts included submissions to
    Approved+locked -- a period with a Draft (or any non-Approved)
    submission is treated as though nothing was entered for that month, not
    as an error.

    An auto-synthesized FY total for a raw monthly field reuses that field's
    own field_id (see synthesize_automatic_fy_totals), so metric_aliases can
    reference either an explicit calculated field or a raw field's
    auto-total by the same field_id, transparently.

    Returns {} if the site has no active workbook or no published sheets.
    """
    from app.modules.SUBMIT import service as submit_svc

    workbook_id = _site_active_workbook_id(site_id)
    if not workbook_id:
        return {}

    sheet_rows = submit_svc._workbook_sheet_rows(workbook_id)
    if not sheet_rows:
        return {}

    cross_sheet_resolver = submit_svc._CrossSheetResolver(site_id, fy_start_year, sheet_rows)
    months = submit_svc._fy_months(fy_start_year)
    month_keys = [(item["year"], item["month"]) for item in months]

    index = {}
    for workbook_form, form in sheet_rows:
        fields = submit_svc._field_payload(form.current_version_id)
        sections = submit_svc._sections_payload(form.current_version_id)
        monthly_fields = submit_svc.monthly_table_fields(fields, sections)
        explicit_result_fields = [f for f in fields if submit_svc.is_sheet_result_field(f)]
        sheet_result_fields = explicit_result_fields + submit_svc.synthesize_automatic_fy_totals(
            monthly_fields, explicit_result_fields
        )

        periods = ReportingPeriod.query.filter(
            ReportingPeriod.site_id == site_id,
            ReportingPeriod.is_deleted == False,
            tuple_(ReportingPeriod.year, ReportingPeriod.month).in_(month_keys),
        ).all()
        period_by_key = {(period.year, period.month): period for period in periods}
        submissions = Submission.query.filter(
            Submission.site_id == site_id,
            Submission.form_id == form.id,
            Submission.is_deleted == False,
            Submission.reporting_period_id.in_([p.id for p in periods] or [0]),
        ).all()
        submission_by_period = {s.reporting_period_id: s for s in submissions}

        rows = []
        for item in months:
            period = period_by_key.get((item["year"], item["month"]))
            submission = submission_by_period.get(period.id) if period else None
            if submission and not include_unapproved:
                if submission.status != "Approved" or not submission.is_locked:
                    submission = None
            values = submit_svc._submission_values_payload(submission, monthly_fields)
            rows.append({
                **item,
                "submission_id": submission.id if submission else None,
                "values": values,
            })

        resolve_external = cross_sheet_resolver.resolve_external_for(form)
        workbook_values = submit_svc.workbook_values_payload(site_id, form.id, fy_start_year, fields)
        sheet_results = submit_svc._compose_sheet_results(
            sheet_result_fields, monthly_fields, rows,
            resolve_external=resolve_external,
            own_sheet_label=workbook_form.sheet_label or submit_svc.human_sheet_label(form),
            annual_fields=fields,
            workbook_values=workbook_values,
        )
        cross_sheet_resolver.finish_external_for(form, sheet_results)

        for result in sheet_results:
            if result.get("value") is not None:
                index[result["field_id"]] = result["value"]

    return index


# Ids of the two computed_columns that are hand-computed in Python, not
# evaluated via Formula -- see validate_computed_columns's docstring for why.
CROSS_SITE_COMPUTED_COLUMN_IDS = {"pct_contribution_total_ghg", "variation_from_avg_intensity"}

# The one ReportTemplate that must go through compose_cross_site_intensity_report
# instead of pivot_report_data -- its config_json's computed_columns include
# kind="cross_site" entries (no formula_id at all), which pivot_report_data's
# _evaluate_computed_columns doesn't know how to skip; routing it there raises
# KeyError('formula_id'). Matches static/js/reports.js's own
# CROSS_SITE_COMPOSER_TEMPLATE_CODE constant -- keep both in sync.
CROSS_SITE_COMPOSER_TEMPLATE_CODE = "cross_site_ghg_intensity_summary"


def compose_cross_site_intensity_report(site_ids, fy_start_year, config):
    """
    Cross-site GHG intensity summary, replicating
    Combined_JSW_Infra_YTD_2025-26.xlsx's Sheet1: two row_groups (SPT / Non-
    SPT), each rendered exactly like pivot_report_data already renders any
    row_group -- site rows, a group subtotal, and a grand total that sums
    the group subtotals (not every site row independently). The ONLY
    structural difference from pivot_report_data is where flat_index comes
    from: _site_flat_index_for_fy (SUM_MONTHS-correct, built directly from
    each site's real per-sheet formula engine) instead of
    generate_report_data's last-write-wins flat rows.

    Two columns need custom, non-generic scoping that _evaluate_computed_
    columns cannot express, and are computed directly in Python here instead
    of via Formula (config["computed_columns"] entries with kind="cross_site"
    are excluded from the normal per-row Formula pass and filled in below):

    - % Contribution of Total Emissions: every row (SPT AND Non-SPT alike)
      divides its own Total GHG Emission by the SPT group's own Total GHG
      Emission -- never the Non-SPT group's own subtotal, never the grand
      total. This is the source workbook's actual formula (=M{row}/$M$10),
      confirmed by inspection, not a bug.
    - Variation from Average Intensity: the SPT group's own weighted GHG
      Intensity (its Total row's own GHG Intensity, not a mean of the 6 SPT
      sites' individual intensities) minus this row's own GHG Intensity --
      populated ONLY for SPT rows. Non-SPT rows, and every Total/Grand Total
      row, get None here, matching source exactly (a deliberate scoping, not
      an oversight to "fix" toward covering every row).

    The SPT group is identified via row_groups' existing is_reference_base
    flag (already validated to be true for exactly one group) -- not a new
    hardcoded group id, reusing the same mechanism pivot_report_data's own
    validate_row_groups already enforces.

    include_unapproved (default True): whether Draft submissions count. This
    report runs before any site's data is locked/approved, by design.
    """
    include_unapproved = config.get("include_unapproved", True)
    metric_aliases = config.get("metric_aliases") or {}
    all_computed_columns = config.get("computed_columns") or []
    formula_computed_columns = [
        c for c in all_computed_columns
        if c.get("kind") != "cross_site" and c.get("id") not in CROSS_SITE_COMPUTED_COLUMN_IDS
    ]

    allowed_site_ids = set(site_ids or [])
    row_groups_config = [
        {**group, "site_ids": [sid for sid in (group.get("site_ids") or []) if sid in allowed_site_ids]}
        for group in (config.get("row_groups") or [])
    ]

    all_group_site_ids = {sid for group in row_groups_config for sid in group["site_ids"]}
    site_names = {s.id: s.name for s in Site.query.filter(Site.id.in_(all_group_site_ids or [0])).all()}

    flat_index = {}
    for site_id in all_group_site_ids:
        for field_id, value in _site_flat_index_for_fy(site_id, fy_start_year, include_unapproved).items():
            flat_index[(site_id, field_id)] = value

    group_subtotal_metrics = {}
    result_groups = []
    for group in row_groups_config:
        group_id = group["id"]
        site_rows = []
        for site_id in group["site_ids"]:
            own_metrics = {}
            metrics_out = {}
            for metric_key in REPORT_CONTEXT_METRIC_KEYS:
                entry = _find_site_alias_entry(metric_aliases.get(metric_key), site_id)
                value, verified = _resolve_alias_entry_value(entry, flat_index, site_id)
                own_metrics[metric_key] = value
                metrics_out[metric_key] = {"value": value, "verified": verified}
            site_rows.append({
                "site_id": site_id,
                "site_name": site_names.get(site_id),
                "metrics": metrics_out,
                "_own_metrics": own_metrics,
            })

        subtotal_metrics = {}
        for metric_key in REPORT_CONTEXT_METRIC_KEYS:
            values = [
                r["metrics"][metric_key]["value"]
                for r in site_rows
                if r["metrics"][metric_key]["value"] is not None
            ]
            subtotal_metrics[metric_key] = sum(values) if values else None

        group_subtotal_metrics[group_id] = subtotal_metrics
        result_groups.append({
            "id": group_id,
            "label": group.get("label"),
            "subtotal_label": group.get("subtotal_label"),
            "is_reference_base": bool(group.get("is_reference_base")),
            "include_in_grand_total": bool(group.get("include_in_grand_total")),
            "suppress_own_subtotal": bool(group.get("suppress_own_subtotal")),
            "site_rows": site_rows,
        })

    group_subtotal_names = {
        f"{group_id}__{metric_key}": value
        for group_id, metrics in group_subtotal_metrics.items()
        for metric_key, value in metrics.items()
        if value is not None
    }

    for group_result in result_groups:
        for site_row in group_result["site_rows"]:
            own_metrics = site_row.pop("_own_metrics")
            site_row["computed"] = _evaluate_computed_columns(
                formula_computed_columns, metric_aliases, flat_index,
                site_row["site_id"], own_metrics, group_subtotal_names,
            )
        group_result["subtotal"] = {
            "label": group_result.pop("subtotal_label"),
            "metrics": group_subtotal_metrics[group_result["id"]],
            "computed": _evaluate_computed_columns(
                formula_computed_columns, metric_aliases, flat_index,
                None, group_subtotal_metrics[group_result["id"]], group_subtotal_names,
            ),
        }

    included_groups = [g for g in result_groups if g["include_in_grand_total"]]
    grand_total_metrics = {}
    for metric_key in REPORT_CONTEXT_METRIC_KEYS:
        values = [
            group_subtotal_metrics[g["id"]][metric_key]
            for g in included_groups
            if group_subtotal_metrics[g["id"]][metric_key] is not None
        ]
        grand_total_metrics[metric_key] = sum(values) if values else None

    grand_total_computed = _evaluate_computed_columns(
        formula_computed_columns, metric_aliases, flat_index, None, grand_total_metrics, group_subtotal_names,
    )

    # --- % Contribution and Variation: custom, SPT-group-targeted ---
    spt_group = next((g for g in result_groups if g["is_reference_base"]), None)
    spt_total_ghg = None
    spt_total_intensity = None
    if spt_group is not None:
        spt_total_ghg = (spt_group["subtotal"]["computed"].get("total_ghg_emission") or {}).get("value")
        spt_total_intensity = (spt_group["subtotal"]["computed"].get("ghg_intensity") or {}).get("value")

    def _blank_cross_site_cell():
        return {"value": None, "source": "cross_site", "verified": None, "error": None}

    for group_result in result_groups:
        is_spt_group = group_result["is_reference_base"]
        for site_row in group_result["site_rows"]:
            row_ghg = (site_row["computed"].get("total_ghg_emission") or {}).get("value")
            pct_cell = _blank_cross_site_cell()
            if row_ghg is not None and spt_total_ghg not in (None, 0):
                pct_cell["value"] = row_ghg / spt_total_ghg * 100
            site_row["computed"]["pct_contribution_total_ghg"] = pct_cell

            variation_cell = _blank_cross_site_cell()
            if is_spt_group:
                row_intensity = (site_row["computed"].get("ghg_intensity") or {}).get("value")
                if spt_total_intensity is not None and row_intensity is not None:
                    variation_cell["value"] = spt_total_intensity - row_intensity
            site_row["computed"]["variation_from_avg_intensity"] = variation_cell

        # Neither cross-site column is meaningful for a Total row itself.
        group_result["subtotal"]["computed"]["pct_contribution_total_ghg"] = _blank_cross_site_cell()
        group_result["subtotal"]["computed"]["variation_from_avg_intensity"] = _blank_cross_site_cell()

    grand_total_computed["pct_contribution_total_ghg"] = _blank_cross_site_cell()
    grand_total_computed["variation_from_avg_intensity"] = _blank_cross_site_cell()

    return {
        "fy_start_year": fy_start_year,
        "include_unapproved": include_unapproved,
        "row_groups": result_groups,
        "grand_total": {
            "metrics": grand_total_metrics,
            "computed": grand_total_computed,
        },
    }


# Display order for pivot-sheet metric columns -- REPORT_CONTEXT_METRIC_KEYS
# itself is a set (no ordering guarantee), so this is a separate, derived
# ordering, not a redefinition of the canonical vocabulary. Public (no leading
# underscore) since RPTBLD/views.py's /api/canonical-metrics route reuses it
# so the frontend never hardcodes this list a third time.
METRIC_KEY_DISPLAY_ORDER = (
    "cargo", "energy_elec", "energy_fossil", "scope1", "scope2", "total_ghg",
    "power_specific", "diesel_specific", "petrol_specific", "ifo_specific",
    "power_consumption_mwh", "diesel_consumption_kl",
)

# Matches the real workbook's Q2:T2 "Specific Consumptions" banner group.
_SPECIFIC_CONSUMPTION_KEYS = ("power_specific", "diesel_specific", "petrol_specific", "ifo_specific")

# The cross-site composer's real Sheet1 column order and header text --
# the single source of truth for BOTH the web preview and the Excel
# export. Metric/computed columns are genuinely interleaved here (energy
# intensity sits right after its raw energy column, etc.), not grouped
# "all metrics then all computed" the way a generic pivot template's
# columns are -- that's the actual shape of the reference workbook this
# template replicates. Public (no leading underscore) since
# RPTBLD/views.py's /api/canonical-metrics route serves this to the
# frontend too (static/js/reports.js fetches it instead of hardcoding its
# own copy), so the two can't drift out of sync the way they did before.
CROSS_SITE_SHEET1_COLUMNS = (
    {"kind": "metric", "id": "cargo", "label": "Cargo Handled (MT)"},
    {"kind": "metric", "id": "energy_elec", "label": "Energy (GJ) - Electrical"},
    {"kind": "computed", "id": "energy_intensity_electrical", "label": "Energy Intensity Electrical (KJ/MT)"},
    {"kind": "metric", "id": "energy_fossil", "label": "Energy (GJ) - Fossil Fuels"},
    {"kind": "computed", "id": "energy_intensity_fossil", "label": "Energy Intensity Fossil Fuels (KJ/MT)"},
    {"kind": "computed", "id": "energy_total_gj", "label": "Energy (GJ)"},
    {"kind": "computed", "id": "energy_intensity_total", "label": "Energy Intensity (000' GJ/Mn MT)"},
    {"kind": "metric", "id": "scope1", "label": "GHG Emission Scope-1 (tCO2e)"},
    {"kind": "computed", "id": "scope1_intensity", "label": "Scope-1 GHG Intensity (KgCO2e/MT)"},
    {"kind": "metric", "id": "scope2", "label": "GHG Emission Scope-2 (tCO2e)"},
    {"kind": "computed", "id": "scope2_intensity", "label": "Scope-2 GHG Intensity (KgCO2e/MT)"},
    {"kind": "computed", "id": "total_ghg_emission", "label": "Total GHG Emission (tCO2e)"},
    {"kind": "computed", "id": "pct_contribution_total_ghg", "label": "% Contribution of Total Emissions"},
    {"kind": "computed", "id": "ghg_intensity", "label": "GHG Intensity (KgCO2e/MT)"},
    {"kind": "computed", "id": "variation_from_avg_intensity", "label": "Variation from Average Intensity"},
    {"kind": "computed", "id": "power_specific_ratio", "label": "Electrical Power (MWH/MnT)"},
    {"kind": "computed", "id": "diesel_specific_ratio", "label": "Diesel (KL/MnT)"},
    {"kind": "computed", "id": "petrol_specific_ratio", "label": "Petrol (KL/MnT)"},
    {"kind": "computed", "id": "ifo_specific_ratio", "label": "IFO/HFHSD (KL/MnT)"},
    {"kind": "metric", "id": "power_consumption_mwh", "label": "Power Consumption MWH"},
    {"kind": "metric", "id": "diesel_consumption_kl", "label": "Diesel Consumption KL"},
)

# Contiguous in CROSS_SITE_SHEET1_COLUMNS' own order -- the Sheet1-mode
# equivalent of _SPECIFIC_CONSUMPTION_KEYS (which names the *raw* metric
# keys used by the generic pivot sheet's banner; Sheet1 only ever shows
# these as their *_ratio computed columns).
_SHEET1_SPECIFIC_CONSUMPTION_RATIO_IDS = (
    "power_specific_ratio", "diesel_specific_ratio", "petrol_specific_ratio", "ifo_specific_ratio",
)


def _is_cross_site_sheet1_config(config):
    """Mirrors reports.js's isCrossSiteSheet1Config exactly -- detects by
    config shape (a kind="cross_site" computed column, or the grand-total
    label unique to this template), not by template code, since this
    helper only ever sees a config_json, never the owning ReportTemplate."""
    computed_columns = (config or {}).get("computed_columns") or []
    return (
        any(c.get("kind") == "cross_site" for c in computed_columns)
        or (config or {}).get("grand_total_label") == "Total All Locations (incl. Non SPT)"
    )


def _pivot_column_specs(config, active_metric_keys, computed_col_ids, computed_col_labels):
    """Single ordered list of {kind, id, label} driving both header cells
    and per-row cell writing -- mirrors reports.js's buildPivotColumnSpecs.
    Sheet1 mode filters the fixed CROSS_SITE_SHEET1_COLUMNS order down to
    whatever's actually configured (some active metric keys -- e.g. the
    raw power_specific/diesel_specific/petrol_specific/ifo_specific --
    deliberately have no column of their own here, only their *_ratio
    computed sibling does, matching the reference sheet exactly). Every
    other template keeps the existing "all metrics, then all computed"
    grouping, unchanged."""
    if _is_cross_site_sheet1_config(config):
        metric_set = set(active_metric_keys)
        computed_set = set(computed_col_ids)

        def _present(col):
            return col["id"] in metric_set if col["kind"] == "metric" else col["id"] in computed_set

        return [col for col in CROSS_SITE_SHEET1_COLUMNS if _present(col)]

    computed_label_by_id = dict(zip(computed_col_ids, computed_col_labels))
    return (
        [{"kind": "metric", "id": k, "label": k.replace("_", " ").title()} for k in active_metric_keys]
        + [{"kind": "computed", "id": cid, "label": computed_label_by_id[cid]} for cid in computed_col_ids]
    )


def _write_pivot_sheet(wb, pivot, config, header_font, bold_font, regular_font, navy_fill, gray_fill, amber_fill, thin_border, sheet_title="Report Data (Pivot)"):
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True

    metric_aliases = config.get("metric_aliases") or {}
    computed_columns = config.get("computed_columns") or []

    # Skip entirely-unconfigured metrics -- no empty columns for metrics nobody's using.
    active_metric_keys = [k for k in METRIC_KEY_DISPLAY_ORDER if metric_aliases.get(k)]
    computed_col_ids = [c["id"] for c in computed_columns]
    computed_col_labels = [c.get("label") or c["id"] for c in computed_columns]

    is_sheet1 = _is_cross_site_sheet1_config(config)
    column_specs = _pivot_column_specs(config, active_metric_keys, computed_col_ids, computed_col_labels)

    # Sheet1 mode merges the generic "Row Group" + "Site" pair into a single
    # "Location" column, matching the reference workbook (which has no
    # separate per-row group column -- SPT/Non-SPT is a visual section, not
    # a cell value). Every other template keeps its existing 2-column header.
    fixed_headers = ["Location"] if is_sheet1 else ["Row Group", "Site"]
    all_headers = fixed_headers + [col["label"] for col in column_specs]
    total_cols = len(all_headers)
    data_col_start = len(fixed_headers) + 1

    # Row 1: banner over whichever specific-consumption columns are actually
    # present -- not a hardcoded fixed range. Sheet1 mode only ever shows
    # these as their *_ratio computed columns, never the raw metric.
    specific_ids = list(_SHEET1_SPECIFIC_CONSUMPTION_RATIO_IDS if is_sheet1 else _SPECIFIC_CONSUMPTION_KEYS)
    specific_kind = "computed" if is_sheet1 else "metric"
    spec_positions = {
        col["id"]: data_col_start + i
        for i, col in enumerate(column_specs)
        if col["kind"] == specific_kind
    }
    specific_present = [k for k in specific_ids if k in spec_positions]
    if specific_present:
        start_col = spec_positions[specific_present[0]]
        end_col = spec_positions[specific_present[-1]]
        banner = ws.cell(row=1, column=start_col, value="Specific Consumptions")
        banner.font = header_font
        banner.fill = navy_fill
        banner.alignment = Alignment(horizontal="center", vertical="center")
        if end_col > start_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

    # Row 2: column headers.
    for col_num, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=col_num, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    def _write_data_columns(row_idx, row_metrics, row_computed, verified_flagging):
        for offset, col in enumerate(column_specs):
            col_num = data_col_start + offset
            if col["kind"] == "metric":
                if verified_flagging:
                    cell_info = row_metrics[col["id"]]
                    value, verified = cell_info["value"], cell_info["verified"]
                else:
                    # Aggregates (subtotal/grand total) have no single
                    # "verified" flag -- never amber-flagged.
                    value, verified = row_metrics[col["id"]], True
                cell = ws.cell(row=row_idx, column=col_num, value=value)
                cell.border = thin_border
                if value is not None:
                    cell.alignment = Alignment(horizontal="right")
                if verified_flagging and not verified:
                    cell.fill = amber_fill
            else:
                cell_info = row_computed.get(col["id"]) or {}
                error = cell_info.get("error")
                cell = ws.cell(row=row_idx, column=col_num)
                cell.border = thin_border
                if error:
                    # Short, single-line error text in the cell -- not blank, not a stack trace.
                    cell.value = error
                    cell.alignment = Alignment(horizontal="left")
                else:
                    value = cell_info.get("value")
                    cell.value = value
                    if value is not None:
                        cell.alignment = Alignment(horizontal="right")
                # Only an override cell's own verified=False is flagged -- a
                # plain "computed" cell's verified is always None and has no meaning.
                if cell_info.get("source") == "override" and cell_info.get("verified") is False:
                    cell.fill = amber_fill

    def _write_fixed_columns(row_idx, group_label, location_label):
        if is_sheet1:
            ws.cell(row=row_idx, column=1, value=location_label)
        else:
            ws.cell(row=row_idx, column=1, value=group_label)
            ws.cell(row=row_idx, column=2, value=location_label)

    row_idx = 3
    for group in pivot["row_groups"]:
        for site_row in group["site_rows"]:
            _write_fixed_columns(row_idx, group.get("label"), site_row.get("site_name"))
            for col_num in range(1, len(fixed_headers) + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = regular_font
                cell.border = thin_border
            _write_data_columns(row_idx, site_row["metrics"], site_row["computed"], verified_flagging=True)
            row_idx += 1

        # suppress_own_subtotal (per-group config flag, mirrors reports.js's
        # renderPivotPreviewTable): this group's own subtotal is never its
        # own row -- its label is really the grand total's label (e.g.
        # "Total All Locations (incl. Non SPT)" IS the grand total, not an
        # independent sum of just this group's rows). The grand-total row
        # written below carries the correct combined value under that label.
        if not group.get("suppress_own_subtotal"):
            subtotal = group["subtotal"]
            _write_fixed_columns(row_idx, group.get("label"), subtotal.get("label"))
            _write_data_columns(row_idx, subtotal["metrics"], subtotal["computed"], verified_flagging=False)
            for col_num in range(1, total_cols + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = bold_font
                cell.fill = gray_fill
                cell.border = thin_border
            row_idx += 1

    # Grand-total label: no such key exists anywhere in phase 2's config_json
    # schema yet, so this uses a fixed string rather than inventing new config
    # schema in this phase -- flagged here as a follow-up if templates need
    # this configurable later.
    grand_total_label = config.get("grand_total_label") or "All Locations"
    if is_sheet1:
        ws.cell(row=row_idx, column=1, value=grand_total_label)
    else:
        ws.cell(row=row_idx, column=2, value=grand_total_label)
    _write_data_columns(row_idx, pivot["grand_total"]["metrics"], pivot["grand_total"]["computed"], verified_flagging=False)
    for col_num in range(1, total_cols + 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.font = header_font
        cell.fill = navy_fill
        cell.border = thin_border

    # Auto column sizing, same approach as the flat "Report Data" sheet.
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)


def export_report_to_excel(template_id, user_id):
    """
    Generates report and writes to formatted Excel sheet.
    """
    t = get_report_template(template_id)
    if not t:
        raise ValueError("Report template not found.")

    # The cross-site template has its own dedicated grouped sheet (below) --
    # the flat, ungrouped "Report Data" dump is generic/shared scaffolding
    # for every OTHER RPTBLD template (kept for those), but is redundant
    # leftover here: the reference workbook it replicates is a single data
    # sheet, and generate_report_data's flat_index is also the same last-
    # write-wins path the cross-site composer was built specifically to
    # bypass -- no reason to still query and render it for this template.
    is_cross_site = t.code == CROSS_SITE_COMPOSER_TEMPLATE_CODE
    data = [] if is_cross_site else generate_report_data(template_id, user_id)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # 1. Cover Sheet
    ws_cover = wb.create_sheet(title="Overview")
    ws_cover.views.sheetView[0].showGridLines = True

    # Fonts & Fills
    title_font = Font(name="Segoe UI", size=16, bold=True, color="1E293B")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="334155")
    regular_font = Font(name="Segoe UI", size=10, color="334155")

    navy_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
    gray_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    # Header Title
    ws_cover["B2"] = t.name
    ws_cover["B2"].font = title_font

    ws_cover["B3"] = "Digital GHG Inventory environmental reporting sheet."
    ws_cover["B3"].font = Font(name="Segoe UI", size=10, italic=True, color="64748B")

    # Metadata Block
    ws_cover["B5"] = "Report Template Metadata"
    ws_cover["B5"].font = bold_font
    ws_cover.merge_cells("B5:C5")

    metadata_rows = [
        ("Template Code", t.code),
        ("Description", t.description or "No description provided."),
        ("Scope Type", t.scope_type.upper()),
        ("Exported At", datetime.now().strftime("%d %b %Y, %I:%M %p")),
    ]

    curr_row = 6
    for k, v in metadata_rows:
        ws_cover.cell(row=curr_row, column=2, value=k).font = bold_font
        ws_cover.cell(row=curr_row, column=2).fill = gray_fill
        ws_cover.cell(row=curr_row, column=2).border = thin_border

        ws_cover.cell(row=curr_row, column=3, value=str(v)).font = regular_font
        ws_cover.cell(row=curr_row, column=3).border = thin_border
        curr_row += 1

    ws_cover.column_dimensions['B'].width = 20
    ws_cover.column_dimensions['C'].width = 50

    # 2. Data Sheet -- skipped entirely for the cross-site template (see note above).
    if not is_cross_site:
        ws_data = wb.create_sheet(title="Report Data")
        ws_data.views.sheetView[0].showGridLines = True

        headers = ["Period", "Site Name", "Form Name", "Field Code", "Field Name", "Field Type", "Value", "Unit"]
        for col_num, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_num, value=h)
            cell.font = header_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row_idx, r in enumerate(data, 2):
            ws_data.cell(row=row_idx, column=1, value=r["period_label"]).font = regular_font
            ws_data.cell(row=row_idx, column=2, value=r["site_name"]).font = regular_font
            ws_data.cell(row=row_idx, column=3, value=r["form_name"]).font = regular_font
            ws_data.cell(row=row_idx, column=4, value=r["field_code"]).font = regular_font
            ws_data.cell(row=row_idx, column=5, value=r["field_name"]).font = regular_font
            ws_data.cell(row=row_idx, column=6, value=r["field_type"]).font = regular_font

            # Value alignment
            val_cell = ws_data.cell(row=row_idx, column=7, value=r["value"])
            val_cell.font = regular_font
            if isinstance(r["value"], (int, float)):
                val_cell.alignment = Alignment(horizontal="right")
            else:
                val_cell.alignment = Alignment(horizontal="left")

            ws_data.cell(row=row_idx, column=8, value=r["unit"]).font = regular_font

            for col_idx in range(1, 9):
                ws_data.cell(row=row_idx, column=col_idx).border = thin_border

        # Auto column sizing
        for col in ws_data.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 3. Pivot Sheet (row groups / metric aliasing / computed columns) --
    # additive only, the flat "Report Data" sheet above is untouched (for
    # templates that still have one).
    config = t.config_json or {}
    if config.get("row_groups"):
        amber_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # Amber 100
        if is_cross_site:
            # This template's computed_columns include kind="cross_site"
            # entries pivot_report_data's _evaluate_computed_columns can't
            # handle (no formula_id at all -- raises KeyError). It also needs
            # the SUM_MONTHS-correct per-site flat index, not
            # generate_report_data's last-write-wins one. Same scoping as
            # the /cross-site-preview endpoint: restrict row_groups' site_ids
            # to what this user can actually see, then let the composer pick
            # the FY that actually has data rather than guessing today's.
            allowed_site_ids, _is_global = _get_user_allowed_sites(user_id, "report")
            row_groups = [
                {**group, "site_ids": [sid for sid in (group.get("site_ids") or []) if sid in allowed_site_ids]}
                for group in (config.get("row_groups") or [])
            ]
            scoped_config = {**config, "row_groups": row_groups}
            site_ids = [sid for group in row_groups for sid in group["site_ids"]]
            fy_start_year = latest_fy_start_year_with_data(site_ids)
            pivot = compose_cross_site_intensity_report(site_ids, fy_start_year, scoped_config)
        else:
            pivot = pivot_report_data(template_id, user_id)
        # "Sheet1" specifically for the cross-site template, mirroring the
        # reference workbook's own single data sheet name -- other
        # row_groups-configured templates keep the generic default name.
        pivot_sheet_title = "Sheet1" if is_cross_site else "Report Data (Pivot)"
        _write_pivot_sheet(wb, pivot, config, header_font, bold_font, regular_font, navy_fill, gray_fill, amber_fill, thin_border, sheet_title=pivot_sheet_title)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()

def get_missing_submissions(user_id):
    """
    Analyzes site, reporting period and applicable forms to track missing sheets.
    """
    allowed_site_ids, is_global = _get_user_allowed_sites(user_id, "submission")

    if not allowed_site_ids:
        return []

    # Get all active open periods
    periods = ReportingPeriod.query.filter(
        ReportingPeriod.site_id.in_(list(allowed_site_ids)),
        ReportingPeriod.status == "OPEN",
        ReportingPeriod.is_deleted == False
    ).order_by(ReportingPeriod.year.desc(), ReportingPeriod.month.desc()).all()

    # Load all published forms
    published_forms = Form.query.filter_by(is_deleted=False).filter(Form.current_version_id.is_not(None)).all()

    sites_map = {s.id: s for s in Site.query.filter_by(is_deleted=False).all()}
    from app.modules.SUBMIT.service import format_period_label

    period_ids = [p.id for p in periods]
    form_ids = [f.id for f in published_forms]

    # Batch-fetch which (site_id, form_id) pairs are assigned via an active
    # workbook -- one query instead of one per (period, form) pair.
    assigned_pairs = set(
        db.session.query(WorkbookSite.site_id, WorkbookForm.form_id)
        .join(Workbook, Workbook.id == WorkbookSite.workbook_id)
        .join(WorkbookForm, WorkbookForm.workbook_id == Workbook.id)
        .filter(
            WorkbookForm.form_id.in_(form_ids or [0]),
            WorkbookSite.site_id.in_(list(allowed_site_ids)),
            Workbook.is_active == True,
        )
        .all()
    )

    # Batch-fetch every existing submission for these periods/forms -- one
    # query instead of one per (period, form) pair.
    existing_submissions = Submission.query.filter(
        Submission.reporting_period_id.in_(period_ids or [0]),
        Submission.form_id.in_(form_ids or [0]),
        Submission.is_deleted == False,
    ).all()
    submissions_by_key = {
        (sub.site_id, sub.form_id, sub.reporting_period_id): sub
        for sub in existing_submissions
    }

    missing_list = []

    for p in periods:
        site = sites_map.get(p.site_id)
        if not site:
            continue

        period_label = format_period_label(p.year, p.month)

        # Check form applicability using WorkbookSite as authoritative source
        for f in published_forms:
            if (p.site_id, f.id) not in assigned_pairs:
                continue

            sub = submissions_by_key.get((p.site_id, f.id, p.id))

            status_desc = "Not Started"
            sub_id = None
            if sub:
                sub_id = sub.id
                status_desc = sub.status

            # If not Approved, it is "missing" (either Not Started, Draft or In Review)
            is_missing = not sub or sub.status != "Approved"

            missing_list.append({
                "site_id": site.id,
                "site_name": site.name,
                "form_id": f.id,
                "form_name": human_sheet_label(f),
                "period_id": p.id,
                "period_label": period_label,
                "submission_id": sub_id,
                "status": status_desc,
                "is_missing": is_missing
            })

    # Sort: Period desc, Site asc, Form Name asc
    missing_list.sort(key=lambda x: (x["period_label"], x["site_name"], x["form_name"]))
    return missing_list
